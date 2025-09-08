import re
import os
import csv
import ast
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from typing import Optional, Tuple
from tqdm import tqdm
from openpyxl import load_workbook
from collections import OrderedDict
import xml.etree.ElementTree as ET

##age_rectify
def rectify_age_meta(
                     gse_gsm_out_csv: str,
                     gse_out_csv:str,
                     mouse_right_csv: str = None,
                     human_right_csv: str = None) -> pd.DataFrame:
    
    # 1. 读取正确文件
    # --------------------------------------------------
    if mouse_right_csv is not None:
        mouse_right = pd.read_csv(mouse_right_csv, sep=',')
        mouse_right.rename(columns={'week_old': 'age_final'}, inplace=True)
        
    if human_right_csv is not None:
        human_right = pd.read_csv(human_right_csv, sep=',')
        human_right.rename(columns={'year_old': 'age_final'}, inplace=True)
        
    # --------------------------------------------------
    # 2. 合并正确数据
    # --------------------------------------------------
    if mouse_right_csv is not None and human_right_csv is not None:
        age_df = pd.concat([mouse_right, human_right], ignore_index=True)
    elif mouse_right_csv is not None:
        age_df = mouse_right
    elif human_right_csv is not None:
        age_df = human_right
    age_df['age_final'] = age_df['age_final'].fillna('Unknown')
    
    # --------------------------------------------------
    # 3. 构造 gse 级 age（当该 GSE 内所有 gsm 不含“GSE”且 age 唯一时）
    # --------------------------------------------------
    gse_ids = age_df['gse'].unique()
    tmp_frames = []

    for i in tqdm(range(len(gse_ids))) :# for gse in gse_ids:
        sub = age_df[age_df['gse'] == gse_ids[i]]
        if (~sub['gsm'].str.contains('GSE')).all():
            uniq_age = sub['age_final'].unique()

            if len(uniq_age) == 1:
                row = sub.iloc[[0]].copy()
                row['gsm'] = gse_ids[i]
                tmp_frames.append(row)
            else:
                row = sub.iloc[[0]].copy()
                row['gsm'] = gse_ids[i]
                age_str = ','.join(map(str, uniq_age))
                row['age_final'] = age_str
                tmp_frames.append(row) 

    gse_age = pd.concat(tmp_frames, ignore_index=True) if tmp_frames else pd.DataFrame()
    gse_age.rename(columns={'Label': 'age_group'}, inplace=True)
    age_df.rename(columns={'Label': 'age_group'}, inplace=True)


    # 确保输出目录存在
    os.makedirs(os.path.dirname(gse_gsm_out_csv), exist_ok=True)
    age_df.to_csv(gse_gsm_out_csv, index=False)
    gse_age.to_csv(gse_out_csv, index= False)
    
    return age_df, gse_age


def rectify_sex_meta(in_csv: str,
                     corr_name: str ,
                     sex_gse_gsm_out_csv: str ,
                     sex_gse_out_csv: str ) -> pd.DataFrame:
    # -------------------- 0. 工具函数 --------------------
    def _check_both_words(text, w1, w2):
        return bool(re.search(rf'.*{w1}.*{w2}.*', text))

    def _only_male_no_female(text):
        return bool(re.match(r'^(?!.*female).*male.*$', text))

    def _letters_sex(text):
        letters = set(filter(str.isalpha, str(text).lower()))
        if letters == {'m'}:
            return 'male'
        elif letters == {'f'}:
            return 'female'
        elif {'m', 'f'} <= letters and 'female' not in str(text).lower():
            return 'female;male'
        elif {'m', 'f'} <= letters:
            return 'female'
        else:
            return 'none'

    # -------------------- 1. 读入数据 --------------------
    gse_gsm_sex = pd.read_csv(in_csv, sep=',')
    gse_gsm_sex['char_sex'] = gse_gsm_sex['char_sex'].astype(str).str.lower()

    # -------------------- 2. 建立标准化映射 --------------------
    none_list = [
        '-', '1', '?', '2', 'unknown', 'not applicable', 'not collected',
        'not determined', 'missing', 'none', 'mixed'
    ]
    del_list = [
        'ebf1fl/fl tie2cre +/+,ebf1wt//wt tie2cre +/cre',
        'ebf1fl/fl tie2cre +/cre,ebf1fl/fl tie2cre +/cre',
        'ebf1wt//wt tie2cre +/cre,ebf1wt//wt tie2cre +/cre'
    ]

    sex_keywords = list(gse_gsm_sex['char_sex'].unique())
    sex_rectify = []

    for sex in sex_keywords:
        sex = str(sex).strip().lower()
        if sex in {'nan', 'none', 'nat', 'nat'} or pd.isna(sex):
            sex_rectify.append('none')
        elif sex == 'both':
            sex_rectify.append('female;male')
        elif sex in none_list:
            sex_rectify.append('none')
        elif sex == 'xx':
            sex_rectify.append('female')
        elif sex == 'xy':
            sex_rectify.append('male')
        elif 'female/female' in sex or 'female,female' in sex:
            sex_rectify.append('female')
        elif _check_both_words(sex, 'female', 'male') or _check_both_words(sex, 'male', 'female'):
            sex_rectify.append('female;male')
        elif _only_male_no_female(sex):
            sex_rectify.append('male')
        elif ('man' in sex or 'men' in sex) and ('woman' not in sex and 'women' not in sex):
            sex_rectify.append('male')
        elif sex in del_list:
            sex_rectify.append('none')
        else:
            sex_rectify.append(_letters_sex(sex))

    sex_corr = pd.DataFrame({'char_sex': sex_keywords, 'sex_new': sex_rectify})

    # 几个手动修正
    sex_corr.loc[sex_corr['char_sex'] == 'xy', 'sex_new'] = 'male'
    sex_corr.loc[sex_corr['char_sex'] ==
                 'mixed canine breed, healthy, female, mixed canine breed, healthy, female',
                 'sex_new'] = 'female'
    sex_corr.loc[sex_corr['char_sex'] ==
                 'multiplexed sample: (e18) pregnant female,  8 week old virgin female and p21 post-partum female',
                 'sex_new'] = 'female'
    sex_corr.loc[sex_corr['char_sex'] ==
                 'human female diploid lung fibroblasts',
                 'sex_new'] = 'female'

    # -------------------- 3. 保存对照表 --------------------
    sex_corr.sort_values('sex_new').to_csv(corr_name,
                                           sep=',', index=False)

    # -------------------- 4. 建立 GSE 粒度唯一 sex 表 --------------------
    gse_gsm_sex['char_sex'] = gse_gsm_sex['char_sex'].fillna('none')
    gse_ids = sorted(gse_gsm_sex['gse'].unique())

    tmp_frames = []
    for i in tqdm(range(len(gse_ids))): # for gse in gse_ids:
        df = gse_gsm_sex[gse_gsm_sex['gse'] == gse_ids[i]].copy()
        # 去掉 gsm 含 GSE 且 sex 为 none 的行
        df = df[~((df['gsm'].str.contains('GSE', na=False)))]
        if (~df['gsm'].str.contains('GSE', na=False)).all():
            uniq_sex = df['char_sex'].unique()
            rec = df.iloc[:1].copy()
            rec['gsm'] = gse_ids[i]
            if len(uniq_sex) == 1:
                rec['char_sex'] = uniq_sex[0]
            else:
                rec['char_sex'] = ','.join(map(str, uniq_sex))
            tmp_frames.append(rec)
                

    gse_gse_sex = pd.concat(tmp_frames, ignore_index=True)
    gse_gse_sex = gse_gse_sex[gse_gse_sex['char_sex'] != 'none']

    # -------------------- 5. 回贴到 gsm 粒度 --------------------
    gse_gsm_rectify = gse_gsm_sex.copy()
    for idx, row in gse_gsm_sex.iterrows():
        m = gse_gse_sex[(gse_gse_sex['gse'] == row['gse']) &
                        (gse_gse_sex['gsm'] == row['gsm'])]
        if not m.empty:
            gse_gsm_rectify.iloc[idx] = m.iloc[0]

    # -------------------- 6. 合并标准化结果 --------------------
    final_df = (gse_gsm_rectify
                .merge(sex_corr[['char_sex', 'sex_new']], on='char_sex', how='inner')
                .rename(columns={'sex_new': 'sex_final'})
                [['gse', 'gsm', 'char_sex', 'sex_final']])

    final_df.to_csv( sex_gse_gsm_out_csv, sep=',', index=False)
    gse_sex = final_df[final_df['gse'].str.contains('GSE')]
    gse_sex.to_csv(sex_gse_out_csv, sep=',', index=False)
    return final_df, gse_sex

def rectify_platform_meta(gse_gsm_platform_path:str,
                          platform_level_path:str,
                          gse_gsm_platform_output_csv:str,
                          gse_platform_output_csv:str):
   
    def unique_platforms_key_word(platform_str):
        # 将字符串转换为列表
        platform_list = ast.literal_eval(platform_str)
        # 获取唯一元素，并转换回列表
        unique_list = sorted(set(platform_list))
        return unique_list
   
    def to_lowercase_and_deduplicate(dictionary):
        lowercased_keys = {k.lower(): v for k, v in dictionary.items()}
        unique_keys = {k: lowercased_keys[k] for k in {k.lower() for k in lowercased_keys}}
        return unique_keys

        ##### 专门处理Smart-seq的情况
    def process_list(my_list):
        target_set = {'Smart-seq', 'Smart-seq2','SMARTer'}
        filtered_list = [item for item in my_list if item not in target_set]
        # 添加'Smart'
        filtered_list.append('Smart-seq2')
        return filtered_list

    ###### This function was used for the platform_key_word rectify
    def replace_and_join(keywords_list):
        # keywords_list = keywords_list.lower()
        unique_keywords = list(set(eval(keywords_list)))
        unique_keywords = [element.lower() for element in unique_keywords] 
        new_list = []
        for platform in unique_keywords:
            if platform in platform_dict:
                new_list.append(platform_dict[platform])
            else:
                new_list.append(platform)  # 如果不在映射中，直接添加
        # 去重并拼接
        unique_platforms = sorted(list(set(new_list)))
        if len(unique_platforms) ==1:
            return ';'.join(unique_platforms)
        elif len(unique_platforms) >1  and  (not set(unique_platforms) & set(unique_platform_l2)):
            return ';'.join(sorted(unique_platforms))
        elif len(unique_platforms) >1 and set(unique_platforms).issubset(set(unique_platform_l2)):  ## 注意 and 和 &的区别
            if '10X Genomics' in unique_platforms:
                unique_platforms = ['10X Genomics']
                return ';'.join(unique_platforms)
            elif 'CITE-seq' in unique_platforms:
                unique_platforms.remove('CITE-seq')
                unique_platforms = sorted(unique_platforms)
                return ';'.join(unique_platforms)
            elif len(unique_platforms)>1 and ('Smart-seq' in unique_platforms  or 'Smart-seq2' in unique_platforms or 'SMARTer' in unique_platforms):
                result = process_list(unique_platforms)   ## 同时出现Smart-seq 和 Smart-seq2, 删除 Smart-seq
                unique_platforms = sorted(result)
                return ';'.join(unique_platforms)
            else:
                return ';'.join(unique_platforms)
        else:
            common_elements = [item for item in unique_platforms if item in set(unique_platform_l2)]
            common_elements = sorted(common_elements)
            # unique_platforms = common_elements
            if '10X Genomics' in common_elements:
                unique_platforms = ['10X Genomics']
                return ';'.join(unique_platforms)
            elif len(common_elements)>1  and ('CITE-seq' in common_elements):
                common_elements.remove('CITE-seq')
                unique_platforms = sorted(common_elements)
                return ';'.join(unique_platforms)
            elif len(common_elements)>1 and ('Smart-seq' in common_elements  or 'Smart-seq2' in common_elements or 'SMARTer' in common_elements ):
                result = process_list(common_elements)   ## 同时出现Smart-seq 和 Smart-seq2, 删除 Smart-seq
                unique_platforms = sorted(result)
                return ';'.join(unique_platforms)
            elif len(common_elements) >=1:
                return ';'.join(common_elements)
            else:
                return ';'.join(unique_platforms)

    gse_gsm_platform = pd.read_csv(gse_gsm_platform_path, sep=',')  ### 989012 rows 
    gse_gsm_platform = gse_gsm_platform.dropna(subset=['platform_key_word']).reset_index(drop=True)    ### 0 rows
    gse_gsm_platform = gse_gsm_platform[gse_gsm_platform['platform_key_word'] != '']   ### 0 rows
    gse_gsm_platform['platform_key_word'] = gse_gsm_platform['platform_key_word'].str.lower()

    platform_level = pd.read_csv(platform_level_path,sep='\t')  # 替换为您的CSV文件路径
    platform_level = platform_level.drop('Remarks',axis =1)
    platform_level = platform_level.drop_duplicates()
    platform_level = platform_level.sort_values(by=['Type','Platform_rename','Platform'], ascending = True)
    platform_level['Platform'] = platform_level['Platform'].str.lower()     #### 为了后面匹配，需要将这个变成小写的先
    platform_level = platform_level.drop_duplicates()    #### 之前的platform会有重复,需要重新去重

    platform_dict = dict(zip(platform_level['Platform'], platform_level['Platform_rename']))  ### 需要将键都改成小写, 不然后面会遗漏

    platform_dict = to_lowercase_and_deduplicate(platform_dict)
    unique_platform_l2 = list(set(platform_dict.values())) 
    unique_platform_l2.remove('droplet/well-based')

    ###### This function was used for the char_platform rectify
    platform_char_new = []
    for row in range(gse_gsm_platform.shape[0]):
        a = gse_gsm_platform['char_platform'][row]
        if (pd.notna(a)):
            if '10x' in str(a).lower():
                b = '10X Genomics'
                platform_char_new.append(b)
            elif 'smart' in str(a).lower():
                b = 'Smart-seq2'
                platform_char_new.append(b)
            else:  ## 不用考虑其他的了，看了下 char_platform 的内容,还有dropseq的都在 platform_key_word中有了，也进行了矫正
                b ='none'
                platform_char_new.append(b)
        else:
            b ='none'
            platform_char_new.append(b)

    gse_gsm_platform['platform_new'] = gse_gsm_platform['platform_key_word'].apply(replace_and_join)
    gse_gsm_platform['platform_new_char'] = platform_char_new    ## 添加新的列
    for row in range(gse_gsm_platform.shape[0]):
        if (gse_gsm_platform['platform_new'][row] == 'none') | (gse_gsm_platform['platform_new'][row] == '') | (pd.isna(gse_gsm_platform['platform_new'][row])):
            gse_gsm_platform.loc[row, 'platform_new'] = gse_gsm_platform.loc[row, 'platform_new_char']
            
    gse_gsm_platform.rename(columns={'platform_new': 'platform_final'}, inplace=True)
    gse_gsm_platform = gse_gsm_platform.drop_duplicates()

    ### 对gsm中为GSE的且platform_final为none的进行补充，看下能否填充
    gse_ids = sorted(list(gse_gsm_platform['gse'].unique()))    # 9246(total), 余 7207 gse_ids( 在上面把none的都去掉，或许会去掉蛮多 GSEid, 所以能不能直接这样去？), 这样其实每个gse在gsm都有一行的
    tmp_frames = []
    for i in tqdm(range(len(gse_ids))):  # ['GSE90697']
        # (~platform_tmp['gsm'].str.contains('GSE')).all():       ##    (any(tissue_tmp['gsm'].str.contains('GSE'))), any 是检查至少有一个 GSE
        platform_tmp = gse_gsm_platform[gse_gsm_platform['gse']==gse_ids[i]].copy()
        platform_a = list(platform_tmp['platform_final'].unique())
        platform_b_inte = list(filter(lambda x: x in unique_platform_l2, platform_a))
        if (len(platform_b_inte)==0):
            platform_a  = [x for x in platform_a if x != 'none']     ## is not
            try:
                platform_a.remove('droplet/well-based')
            except ValueError:
                pass
            if (len(platform_a)==1):
                platform_tmp['platform_final'] = platform_a[0]
                tmp_frames.append(platform_tmp)  # python3.11.4用._append() ; 3.9.12用append()
            elif (len(platform_a)==0):
                platform_tmp['platform_final'] ='none'
                tmp_frames.append(platform_tmp)
            else:
                if ('10X Genomics' in platform_a):
                    platform_tmp['platform_final'] = '10X Genomics'
                    tmp_frames.append(platform_tmp)
                elif ('Smart-seq' in platform_a):
                    platform_tmp['platform_final'] = 'Smart-seq'
                    tmp_frames.append(platform_tmp)
                else:
                    platform_tmp.loc[platform_tmp['gsm'].str.contains('GSE'), 'platform_final'] = platform_a[0]
                    tmp_frames.append(platform_tmp)
        elif (len(platform_b_inte)==1):
            platform_tmp['platform_final'] = platform_b_inte[0]
            tmp_frames.append(platform_tmp)
        else:
            if ('10X Genomics' in platform_b_inte):
                platform_tmp['platform_final'] = '10X Genomics'
                tmp_frames.append(platform_tmp)
            elif ('Smart-seq' in platform_b_inte):
                platform_tmp['platform_final'] = 'Smart-seq'
                tmp_frames.append(platform_tmp)
            else:
                platform_tmp.loc[platform_tmp['gsm'].str.contains('GSE'), 'platform_final'] = platform_b_inte[0]
                tmp_frames.append(platform_tmp)

    gse_gsm_platform_new = pd.concat(tmp_frames, ignore_index=True)

    gse_gsm_platform_new[gse_gsm_platform_new['platform_final']=='none']['gse'].unique()   ## 1331 个gse是none的, 131025 row是none的

    gse_gsm_platform_new[gse_gsm_platform_new['gsm'].str.contains('GSE')]['platform_final'].value_counts()

    ### 再遍历gse_gsm_platform的每一行，将其中不常见的给去掉，如果有smart-seq系列或者10x系列,同时又有C1，很有可能就是 
    gse_gsm_platform_new[gse_gsm_platform_new['gsm'].str.contains('GSE')]['platform_final'].value_counts()   ## 有 6313 个gsm中是GSE的platform信息为none，后面可以去看下其对应的gsm下面有没有信息
    sorted(gse_gsm_platform_new['platform_final'].unique())    #### 有 57 个 unique的 item, 需要对这 57 个item 进行一些批量处理
    gse_gsm_platform_new = gse_gsm_platform_new[['gse', 'gsm',  'char_platform', 'platform_key_word','platform_final']]

    unique_platform_item = sorted(gse_gsm_platform_new['platform_final'].unique())   ## 52 个 item

    none_list = ['truseq','novaseq','mrna-seq','hiseq2000','hiseq2500','hiseq4000','novaseq_sp','droplet/well-based','geo-seq','transplex']  ### hiseq这是一个系列的

    unique_platform_new  = []
    for item in unique_platform_item:
        parts = item.split(';')
        # filtered_parts = [part for part in parts if part.strip() not in none_list]
        filtered_parts = [sub_item for sub_item in parts if sub_item not in none_list]
        if(len(filtered_parts)>=1):
            sorted_parts = sorted(filtered_parts)    ### 就是有 hiseq 的，
            new_item = ';'.join(sorted_parts)
            unique_platform_new.append(new_item)
        else:
            sorted_parts = ['none']
            new_item = ';'.join(sorted_parts)
            unique_platform_new.append(new_item)

    platform_df = pd.DataFrame({'platform_final':unique_platform_item, 'platform_new':unique_platform_new})  ## 新的 51 个 item
    gse_gsm_platform_new = pd.merge(gse_gsm_platform_new, platform_df, on=['platform_final'], how='inner') 
    gse_gsm_platform_new = gse_gsm_platform_new[['gse','gsm','char_platform','platform_key_word','platform_new']]
    gse_gsm_platform_new.rename(columns={'platform_new': 'platform_final'}, inplace=True)  ## 132393 rows为none
    gse_gsm_platform_new.loc[gse_gsm_platform_new['platform_final'].str.contains('Smart-seq2', case=False, na=False), 'platform_final'] = 'Smart-seq2'

    # 找到重复的 gsm 行
    duplicates = gse_gsm_platform_new[gse_gsm_platform_new['gsm'].duplicated(keep=False)]

    # 对于每个重复的 gsm，检查 platform_final 是否一致
    for gsm, group in duplicates.groupby('gsm'):
        # 获取 platform_final 列的唯一值
        platforms = group['platform_final'].unique()
        
        # 如果存在多个不同的值
        if len(platforms) > 1:
            # 过滤掉 'none' 值
            platforms = [platform for platform in platforms if platform.lower() != 'none']
            
            # 如果过滤后还有多个值，则合并这些值
            if len(platforms) > 1:
                combined_platform = ', '.join(sorted(platforms))
            elif len(platforms) == 1:
                combined_platform = platforms[0]
            else:
                combined_platform = None  # 如果过滤后没有值，则设置为 None
            
            # 找到这些行的索引
            indices = group.index
            
            # 更新这些行的 platform_final
            gse_gsm_platform_new.loc[indices, 'platform_final'] = combined_platform
    
    gse_gsm_platform_new.to_csv(gse_gsm_platform_output_csv,sep=',',index=False)

    gse_platform_new = gse_gsm_platform_new[gse_gsm_platform_new['gsm'].str.contains('GSE')]

    gse_platform_new.to_csv(gse_platform_output_csv,sep=',',index=False)

    return gse_gsm_platform_new, gse_platform_new

def rectify_disease_meta(gse_gsm_disease_path:str,
                         disease_type_path:str,
                         disease_name_type_path:str,
                         gse_gsm_disease_output_csv:str,
                         gse_disease_output_csv:str):
    
    def create_dict_from_disease_name_df(df):
        result_dict = {}
        for index, row in df.iterrows():
            word = row['Disease']
            if word.isupper() or (word[1:-1].isupper() and word[-1].islower()):   ### 全部为大写,或者除了收尾都为大写(iPSCs)
                word = word
            else:   
                word = word.lower()
            unique_disease = word
            result_dict[unique_disease] = row['Type']
        return result_dict
    
    
    # 处理 char_disease 列
    def get_char_type(row):
        if pd.isna(row['char_disease']):
            diseases = 'none'
        elif not isinstance(row['char_disease'],str):
            a = str(row['char_disease'])
            diseases = a.lower()
        else:
            diseases = row['char_disease'].lower()
        types = []
        if diseases in disease_type_dict:
            types.append(disease_type_dict[diseases])
        return ','.join(types) if types else 'none'

    # 处理 disease_key_word 列
    def get_disease_type(row):
        # 将字符串转换为列表，并去除重复元素
        try:
            keywords = ast.literal_eval(row['disease_key_word'])
            keywords = list(set(keywords))
            keywords = [element.lower() for element in keywords]
        except (ValueError, SyntaxError):
            keywords = []
        types = []
        for keyword in keywords:
            if keyword in disease_type_dict:
                types.append(disease_type_dict[keyword])
        return ','.join(set(types)) if types else 'none'
    
    def classify_disease(disease):
        print(disease)
        if pd.isna(disease):  # 判断是否为NaN
            return 'none'
        disease_lower = disease.lower()  # 转为小写以便于匹配
        # 遍历normal_keyword，如果匹配则返回'normal'
        if any(keyword in disease_lower for keyword in normal_keyword):
            return '25'
        # 遍历tumor_keyword，如果匹配则返回'tumor'
        elif any(keyword in disease_lower for keyword in tumor_keyword):
            return '19'
        else:
            return 'none'

    
    def get_disease_final(row):
        if row['disease_type'] != 'none':
            return row['disease_type']
        elif row['char_type'] != 'none':
            return row['char_type']
        elif row['char_disease_new'] != 'none':
            return row['char_disease_new']
        else:
            return 'none'
        
    def map_disease_final_to_name(disease_final):
        diseases = disease_final.split(',')
        diseases = list(set(diseases))
        if(diseases==['none']):
            labels = [disease_map[d] for d in diseases if d in disease_map]
        else:
            labels = [disease_map[d] for d in diseases if d in disease_map]  ## 不需要 int(d)
        return ','.join(sorted(labels))
    
    def get_disease_labels(type_list, disease_type_df):
        labels = []
        for type_ in type_list:
            match = disease_type_df[disease_type_df['Type'] == type_]
            if not match.empty:
                labels.append(match['Label'].values[0])
        return ','.join(sorted(labels))


    def calculate_disease_final(row, disease_type_df):
        char_type = row['char_type']
        disease_type = row['disease_type']
        if char_type == 'none' and disease_type == 'none':
            return 'none'
        elif char_type == 'none' and disease_type!='none':
            disease_types = disease_type.split(',')
            return get_disease_labels(disease_types, disease_type_df)
        elif disease_type == 'none' and char_type != 'none':
            char_types = char_type.split(',')
            return get_disease_labels(char_types, disease_type_df)
        else:
            char_types = char_type.split(',')
            disease_types = disease_type.split(',')
            if len(char_types) <= len(disease_types):
                return get_disease_labels(char_types, disease_type_df)
            else:
                return get_disease_labels(disease_types, disease_type_df)

    def fill_column(col):
    # 优先选择没有','且不为None的值
        no_comma_values = [x for x in col if (x != 'None') and (',' not in str(x))]
        with_comma_values = [x for x in col if (x != 'None') and (',' not in str(x))]
        for i in range(len(col)):
            if col[i] == 'None':
                if no_comma_values:
                    col[i] = no_comma_values[0]  # 使用第一个没有逗号的值
                elif with_comma_values:
                    col[i] = with_comma_values[0]  # 使用第一个有逗号的值
        return col
    
    def fill_none_with_value(group):
    # 获取组内第一个非None值
        first_valid = group.dropna()
        if not first_valid.empty:
            return group.fillna(first_valid.iloc[0])
        return group

    gse_gsm_disease = pd.read_csv(gse_gsm_disease_path,sep=',')
    disease_sub = gse_gsm_disease[['gse','gsm','char_disease','disease_key_word']].dropna()
    disease_sub = disease_sub[disease_sub['disease_key_word'] != '[]']

    disease_type = pd.read_csv(disease_type_path ,sep=',',header=None)
    disease_type.columns =['Type','Label']
    disease_type_list = disease_type['Label'].to_list()
    disease_type_list_new = [s.replace('\xa0', ' ') for s in disease_type_list]
    disease_type['Label'] = pd.DataFrame(disease_type_list_new)
    disease_type['Type_l1'] ='Anatomical'
    disease_type.loc[disease_type['Label'].isin(['Cancer','Fetal disease','Genetic disease','Infectious disease','Metabolic disease','Rare disease','Normal']),'Type_l1'] = 'Global'
    disease_type['Type'] = disease_type['Type'].astype(str)
    # disease_type[disease_type['Label'].isin(['Cancer','Fetal disease','Genetic disease','Infectious disease','Metabolic disease','Rare disease'])]['Type_l1'] = 'Global'
    disease_name_type =pd.read_csv(disease_name_type_path,sep='\t')
    disease_sub['char_type'] = ''
    disease_sub['disease_type'] = ''

    disease_type_dict = create_dict_from_disease_name_df(disease_name_type)
    disease_type_dict = {key.lower(): value for key, value in disease_type_dict.items()}

    
    normal_keyword =['normal','healthy','disease-free','disease free','healty','control','uninfected','no-covid','non covid','health','uninjured']
    tumor_keyword =['benign','metastasis','carcinoma','tumor','cancer','adenocarcinoma','leiomyoma','liposarcoma','sarcoma','adenoma']
    gse_gsm_disease['char_disease'] = gse_gsm_disease['char_disease'].str.lower()
    gse_gsm_disease['disease_key_word'] = gse_gsm_disease['disease_key_word'].str.lower()
    gse_gsm_disease['char_type'] = gse_gsm_disease.apply(get_char_type, axis=1)
    gse_gsm_disease['disease_type'] = gse_gsm_disease.apply(get_disease_type, axis=1)
    gse_gsm_disease[(gse_gsm_disease['char_type']=='none') & (gse_gsm_disease['disease_type']=='none')]    ## 727171 rows

    # 使用apply函数来遍历disease列并创建新的disease_new列
    gse_gsm_disease['char_disease_new'] = gse_gsm_disease['char_disease'].apply(classify_disease)
    gse_gsm_disease[(gse_gsm_disease['char_type']=='none') & (gse_gsm_disease['disease_type']=='none') & (gse_gsm_disease['char_disease_new']=='none')] #  717836 rows

    
    gse_gsm_disease['disease_final'] = gse_gsm_disease.apply(get_disease_final, axis=1)
    disease_map = disease_type.set_index('Type')['Label'].to_dict()
    disease_map.update(none='None')

    
    gse_gsm_disease['disease_name'] = gse_gsm_disease['disease_final'].apply(map_disease_final_to_name)
    gse_gsm_disease['Global_D'] ='None'
    gse_gsm_disease['Anatomical_D'] ='None'
    global_disease_names = disease_type[disease_type['Type_l1']=='Global']['Label'].to_list()
    anatomical_disease_names = disease_type[disease_type['Type_l1']=='Anatomical']['Label'].to_list()

    gse_gsm_disease_names_unique = list(gse_gsm_disease['disease_name'].unique())  
    for name in range(len(gse_gsm_disease_names_unique)):
        disease_name = gse_gsm_disease_names_unique[name]
        print(disease_name)
        if (disease_name =='None'):
            gse_gsm_disease.loc[gse_gsm_disease['disease_name']==disease_name,'Global_D'] = 'None'
            gse_gsm_disease.loc[gse_gsm_disease['disease_name']==disease_name,'Anatomical_D'] = 'None'
        else:
            disease_names = list(set(disease_name.split(',')))
            names_global = sorted(list(set(disease_names) & set(global_disease_names))) 
            names_anatomical = sorted(list(set(disease_names) & set(anatomical_disease_names)))
            names_global_label = ','.join(sorted(names_global))
            names_anatomical_label = ','.join(sorted(names_anatomical))
            gse_gsm_disease.loc[gse_gsm_disease['disease_name']==disease_name,'Global_D'] = names_global_label
            gse_gsm_disease.loc[gse_gsm_disease['disease_name']==disease_name,'Anatomical_D'] = names_anatomical_label

    # gse_gsm_disease = gse_gsm_disease[['gse', 'gsm', 'char_type', 'disease_type', 'char_disease_new', 'disease_final', 'disease_name']]
    gse_gsm_disease = gse_gsm_disease[['gse','gsm','char_disease','disease_key_word','Global_D','Anatomical_D']]  ## 'disease_name'
    gse_gsm_disease.rename(columns={'Global_D': 'Global_type'}, inplace=True)
    gse_gsm_disease.rename(columns={'Anatomical_D': 'Anatomical_type'}, inplace=True)

        
    gse_gsm_disease['Global_type'].replace('','None',inplace=True)
    gse_gsm_disease['Anatomical_type'].replace('','None',inplace=True)
    disease_none_gse = list(gse_gsm_disease[(gse_gsm_disease['Global_type']=='None') & (gse_gsm_disease['Anatomical_type']=='None')]['gse'].unique())  ### 这里是Global和Anatomical都没有disease信息的, 717836 rows
  
    
    for gse in disease_none_gse:  # 'GSE100060'
        gse_gsm_disease.loc[gse_gsm_disease['gse']==gse,'Global_type'] = fill_column(gse_gsm_disease[gse_gsm_disease['gse']==gse]['Global_type'].tolist())  # 不知道这个O不OK
        gse_gsm_disease.loc[gse_gsm_disease['gse']==gse,'Anatomical_type'] = fill_column(gse_gsm_disease[gse_gsm_disease['gse']==gse]['Anatomical_type'].tolist())
        #gse_gsm_disease[gse_gsm_disease['gse']==gse]['Global_type'] = fill_column(gse_gsm_disease[gse_gsm_disease['gse']==gse]['Global_type'].tolist())
        #gse_gsm_disease[gse_gsm_disease['gse']==gse]['Anatomical_type'] = fill_column(gse_gsm_disease[gse_gsm_disease['gse']==gse]['Anatomical_type'].tolist()) 

    # gse_gsm_disease[(gse_gsm_disease['Global_type']=='None') & (gse_gsm_disease['Anatomical_type']=='None')]
    
    gse_gsm_disease['Disease_Global_type'] = gse_gsm_disease.groupby('gsm')['Global_type'].transform(fill_none_with_value)
    gse_gsm_disease['Disease_Anatomical_type'] = gse_gsm_disease.groupby('gsm')['Anatomical_type'].transform(fill_none_with_value)
    
    
    gse_gsm_disease.to_csv(gse_gsm_disease_output_csv,sep=',',index=False)

    gse_disease = gse_gsm_disease[gse_gsm_disease['gsm'].str.contains('GSE')]

    gse_disease.to_csv(gse_disease_output_csv,sep=',',index=False)

    return gse_gsm_disease, gse_disease


def rectify_tissue_meta(gse_gsm_tissue_path:str,
                        tissue_level_path:str,
                        cancer_cellline_path:str,
                        gse_gsm_tissue_output_csv:str,
                        gse_tissue_output_csv:str
                        ):
   
   def create_dict_from_tissue_df(df):
      result_dict = {}
      for index, row in df.iterrows():
         # 将Tissue列转换为小写并按逗号分割
         tissues = row['tissue'].split(',')
         processed_words = []
         for word in tissues:
            if word.isupper() or (word[1:-1].isupper() and word[-1].islower()):   ### 全部为大写,或者除了收尾都为大写(iPSCs)
               processed_words.append(word)
            else:   
               processed_words.append(word.lower())
         unique_tissues = set(sorted(processed_words))
         for tissue in unique_tissues:
               result_dict[tissue] = row['tissue_type']
      return result_dict
   

   def process_list(my_list):
      target_set = {'tumor','cancer','h4','u2','carcinoma','del'}  # ‘none' 去掉，因为总会有没tissue的情况
      filtered_list = [item for item in my_list if item not in target_set]
      # filtered_list.append('Smart-seq2')
      return filtered_list

   def match_tissue(tissue_str, tissue_dict):
      if not isinstance(tissue_str,str):
         tissue_str = str(tissue_str)
         tissue_str = tissue_str.lower()
      else:
         tissue_str = tissue_str.lower()
      match_value =[]
      for key, value in tissue_dict.items():
         if key in tissue_str or value in tissue_str:
               match_value.append(value)   ### Tissue这里也有问题的
      match_value = list(sorted(set(match_value)))
      if len(match_value) >=1:
         return ';'.join(match_value)
      else:
         return 'none'

   def replace_and_join_tissue(keywords_list1):  # keywords_list2 这个函数是最后那列的
      unique_keywords1 = list(set(eval(keywords_list1)))
      new_list = []  ### 给个list来确定
      for tissue in unique_keywords1:
         if tissue in tissue_dict_one:
            new_list.append(tissue_dict_one[tissue])
         else:
            new_list.append(tissue)  # 如果不在映射中，直接添加
      # 去重并拼接
      unique_tissues = sorted(list(set(new_list)))
      if len(unique_tissues) ==1:
         return ';'.join(unique_tissues)  ## only 1 tissue
      ## 没有交集
      elif len(unique_tissues)>1 and (not set(unique_tissues) & set(unique_tissue_l1)):
         if len(unique_tissues)>1 and (('tumor' in unique_tissues) or ('cancer' in unique_tissues) or ('h4' in unique_tissues) or ('u2' in unique_tissues) or ('carcinoma' in unique_tissues) or ('del' in unique_tissues)):
            result = process_list(unique_tissues)   ## 同时出现Smart-seq 和 Smart-seq2, 删除 Smart-seq
            unique_tissues_1 = sorted(result)
            if len(unique_tissues_1)==0:
               unique_tissues_1 = ['none']
               return ';'.join(unique_tissues_1)  ## without interaction
            else:
               return ';'.join(unique_tissues_1) 
      ## unique_tissues 都在tissue type中,不会有tumor, h4这些情况, 但这里可能有多个item的情况,这里需要考虑舍弃和保留的问题
      elif len(unique_tissues)>1 and set(unique_tissues).issubset(set(unique_tissue_l1)): 
         if len(unique_tissues)>1 and ('none' in unique_tissues):
            unique_tissues.remove('none')
            return ';'.join(unique_tissues)
         else:
            return ';'.join(unique_tissues)
      ## 这下面的是unique_tissues 不一定都在tissue type中(但一定会有),所以碰到一些奇怪的信息要看下, tissue和platform不一样，这里需要看下common_elements的length,
      else:
         common_elements = [item for item in unique_tissues if item in set(unique_tissue_l1)]
         common_elements = sorted(common_elements) 
         ###### 需要考虑下是否仅有GSM而没有GSE的情况
         if len(common_elements)==1:
            return ';'.join(common_elements) 
         if len(common_elements)>1:
            if len(common_elements)>1 and ('none' in common_elements):
               common_elements.remove('none')
               return ';'.join(common_elements) 
            else:
               return ';'.join(common_elements) 
            
   def replace_none_with_tissue(df):
      new_df = df.copy()  # 创建一个副本，避免修改原DataFrame
      # new_df['tissue_new_char'] = new_df.apply(lambda row: row['tissue_new_char'] if row['tissue_new_char'] != 'none' else row['tissue_new'], axis=1)
      for index, row in new_df.iterrows():
         if pd.isna(row['tissue_new_char']) or  row['tissue_new_char'].lower()=='none':
            new_df.loc[index,'tissue_new_char'] = row['tissue_new']
      return new_df
   
   def assign_tissue_type(df, cancer_cellline, unique_tissue_l1):
    # 定义一个函数来对每一行的tissue_final进行处理
    def check_tissue_type(tissue_final_str):
        tissue_list = [t.strip() for t in tissue_final_str.split(';')]
        if tissue_list==['none']:
            return 'none'
        elif any(t in cancer_cellline for t in tissue_list):
            return 'cancer cellline'
        elif any(t in unique_tissue_l1 for t in tissue_list):
            return 'tissue'
        else:
            return 'cellline'
    df['tissue_type'] = df['tissue_final'].apply(check_tissue_type)
    return df



   gse_gsm_tissue = pd.read_csv(gse_gsm_tissue_path,sep=',')  ## 989012 rows
   gse_gsm_tissue = gse_gsm_tissue.dropna(subset=['tissue_key_word']).reset_index(drop=True)   ### 0 rows (na值), 余 989012 rows
   gse_gsm_tissue = gse_gsm_tissue[gse_gsm_tissue['tissue_key_word'] != '']   ### 0 rows
   gse_gsm_tissue['tissue_key_word'] = gse_gsm_tissue['tissue_key_word'].str.lower()
   tissue_level = pd.read_csv(tissue_level_path,sep='\t',header=None)
   tissue_level.columns = ['tissue','tissue_type'] 
   tissue_level = tissue_level[tissue_level['tissue_type'] != 'carcinoma']


   tissue_dict_one = create_dict_from_tissue_df(tissue_level)   ### 不分开 tissueType,看需要获取, 723 items

   unique_tissue_l1 = sorted(list(set(tissue_dict_one.values())))


   gse_gsm_tissue['tissue_new'] =gse_gsm_tissue['tissue_key_word'].apply(replace_and_join_tissue)  ### tissue_new # 这样会有 '' 字符串
   gse_gsm_tissue['tissue_new_char'] = gse_gsm_tissue['char_tissue'].apply(lambda x: match_tissue(x, tissue_dict_one))
   gse_gsm_tissue[gse_gsm_tissue['tissue_new'] ==''] # 0 rows,  gse_gsm_tissue[pd.isna(gse_gsm_tissue['tissue_new'])] , 这时候还有 140140 行 pd.isna(), 2277 'none'; tissue_new_char列有 636065 'none', 0 rows '', 0 rows None

   result_df = replace_none_with_tissue(gse_gsm_tissue)  # tissue_new_char 中为None 或 none的，看下 tissue_new 中有无，有的话用tissue_new的值替换
   result_df[result_df['tissue_new_char']=='none']   ## 还有部分是 tissue_new 也没有的, 2055( 'none' )+ 122753 (None)= 124808 rows; 0 rows '', 最终需要的是tissue_new_char中的信息
   result_df.rename(columns={'tissue_new_char': 'tissue_final'}, inplace=True)  
   result_df['tissue_final'] = result_df['tissue_final'].fillna('none')   ### 这样就没有 na 值了

   # a = [element for element in unique_tissue if 'bone marrow' in element]
   # b = [element for element in unique_tissue if 'bone;bone marrow' in element]  # 差集中没有bone marrow和marrow同时出现的
   result_df['tissue_final'] = result_df['tissue_final'].str.replace('bone;bone marrow','bone marrow') 
   # result_df = result_df[result_df['tissue_final']!='none']
   # result_df = result_df.dropna(subset=['tissue_final'])     ### 余下 800661 rows有tissue信息，但 188351 rows不需要去掉，没有的就记为none就好

   tissue_none_gse = list(result_df[result_df['tissue_final']=='none']['gse'].unique())
   tissue_none_table_df = result_df[result_df['gse'].isin(tissue_none_gse)]
   ### 下面这步获取gse_gse information的适合爬取的时候GSE page没有爬到信息的时候用，在这里，tissue的不需要用到，可以不用跑下面这步, 但跑一下可以验证
   tmp_list = []
   for i in  tqdm(range(len(tissue_none_gse))):  #   tissue_none_gse
      tissue_tmp = result_df[result_df['gse']==tissue_none_gse[i]].copy()
      tissue_tmp_1 = tissue_tmp[tissue_tmp['tissue_final']!='none']
      tissue_a = list(tissue_tmp_1['tissue_final'].unique())
      if (len(tissue_a)==1):
         tissue_tmp['tissue_final'] = tissue_a[0]
         tmp_list.append(tissue_tmp)
      elif (len(tissue_a)==0):
         tissue_tmp['tissue_final'] = 'none'
         tmp_list.append(tissue_tmp)
      else:
         tissue_a_len = [len(item.split(';')) for item in tissue_a]
         one_indices =  [i for i, x in enumerate(tissue_a_len) if x == 1]   #  if (len(one_indices)>=1)
         if len(one_indices)>=1:
            # tissue_tmp['tissue_final'] = tissue_tmp['tissue_final'].str.replace('none',tissue_a[one_indices[0]]) 
            tissue_tmp.loc[tissue_tmp['tissue_final']=='none','tissue_final'] = tissue_a[one_indices[0]]
            tmp_list.append(tissue_tmp)
         else:
            two_indices =  [i for i, x in enumerate(tissue_a_len) if x == min(tissue_a_len)]
            # tissue_tmp['tissue_final'] = tissue_tmp['tissue_final'].str.replace('none',tissue_a[two_indices[0]]) 
            tissue_tmp.loc[tissue_tmp['tissue_final']=='none','tissue_final'] = tissue_a[two_indices[0]]
            tmp_list.append(tissue_tmp)

   gse_gsm_tissue_none_new = pd.concat(tmp_list,ignore_index=True)
   

   ## gse_gsm_tissue_none_new.loc[gse_gsm_tissue_none_new.duplicated()]  # 找重复的行
   tissue_yes_table_df = result_df[-result_df['gse'].isin(tissue_none_gse)]
   result_df_new = pd.concat([tissue_yes_table_df,gse_gsm_tissue_none_new],ignore_index=True, sort=False)
   result_df_new.index = range(len(result_df_new))  # 只有 366 个gse是none的了

   
   # gse_gsm_tissue_rectify[gse_gsm_tissue_rectify['tissue_final']=='none']  ## 检查下
   result_df_new['tissue_final'] = result_df_new['tissue_final'].str.replace('none;','')  
   result_df_new['tissue_final'] = result_df_new['tissue_final'].str.replace(';none','')
   gse_gsm_tissue_final = result_df_new[['gse','gsm','char_tissue','tissue_key_word','tissue_final']]
   gse_gsm_tissue_final['tissue_final'] = gse_gsm_tissue_final['tissue_final'].str.replace('bone;bone marrow','bone marrow')   ### bone;bone marrow同时出现是有问题的

   ###### add 'tissue_type' column, the cellline or tissue(normal, cancer), how to define? cancer_cellline ? Simplify_new ?
   cancer_cellline = pd.read_csv(cancer_cellline_path,sep='\t',header=None)
   cancer_cellline.columns = ['Cancer cellline']
   
   gse_gsm_tissue_processed = assign_tissue_type(result_df_new, cancer_cellline['Cancer cellline'].to_list(), unique_tissue_l1)

   result_df_new.to_csv(gse_gsm_tissue_output_csv,sep=',',index=False) 

   gse_tissue_df = result_df_new[result_df_new['gsm'].str.contains('GSE')]

   gse_tissue_df.to_csv(gse_tissue_output_csv,sep=',',index=False) 

   return result_df_new, gse_tissue_df

def rectify_treatment_meta(gse_gsm_treat_path:str,
                           drug_list_path:str,
                           gene_key_path:str,
                           control_path:str,
                           cyto_key_path:str,
                           gse_gsm_treatment_output_csv:str,
                           gse_treatment_output_csv:str
                           ):
    
    def remove_parentheses(list_with_tuples):
        new_list = [item[0] for item in list_with_tuples]
        return new_list
    
    def create_sorted_dict(drug, gene, cyto, control):
        data = {
            key: value for value, keys in [
                ('drug', drug),
                ('gene', gene),
                ('cyto', cyto),
                ('control', control)
            ] for key in keys
        }
        sorted_data = OrderedDict(sorted(data.items(), key=lambda x: (x[1], x[0])))
        return sorted_data
    
    def match_treat(treat_str, treatment_level_dict):
        if pd.isna(treat_str):   ### 重新爬的会有 char_treatment 是 nan的情况
            treat_str = ''
        # isinstance函数要注意不能把 str 带进来作为变量
        elif not isinstance(treat_str,str):
            treat_str = str(treat_str)
            treat_str = treat_str.lower()
        else:
            treat_str = treat_str.lower()
        match_value = []
        for key, value in treatment_level_dict.items():
            #print(key)
            #print(value)
            if key in treat_str or value in treat_str:
                match_value.append(value)
            # else:
                # print('None')
        match_value = sorted(list(set(match_value)))
        if len(match_value)>=1:
            return ';'.join(match_value)
        else:
            return 'none'


    def match_treat_key(treat_str, treatment_level_dict):
        if pd.isna(treat_str):   ### 重新爬的会有 char_treatment 是 nan的情况
            treat_str = ''
        # isinstance函数要注意不能把 str 带进来作为变量
        elif not isinstance(treat_str,str):
            treat_str = str(treat_str)
            treat_str = treat_str.lower()
        else:
            treat_str = treat_str.lower()
        match_value = []
        for key, value in treatment_level_dict.items():
            if key in treat_str or value in treat_str:
                if value == 'control':
                    match_value.append(value)
                else:
                    match_value.append(key)
            # else:
                # print('None')
        match_value = sorted(list(set(match_value)))
        if len(match_value)>=1:
            return ';'.join(match_value)
        else:
            return 'none'
        

    # 定义函数来确定treat_final的值
    def get_treat_type(row):
        for col in type_priority_columns:
            if row[col] != 'none':
                return row[col]
        return 'none'

    def get_treat_final(row):
        for col in word_priority_columns:
            if row[col] != 'none':
                return row[col]
        return 'none'
    
    def fill_none_with_value(group):
    # 获取组内第一个非None值
        first_valid = group.dropna()
        if not first_valid.empty:
            return group.fillna(first_valid.iloc[0])
        return group
            
    gse_gsm_treat = pd.read_csv(gse_gsm_treat_path,sep=',')  ## 989012 rows
    gse_gsm_clean = gse_gsm_treat.dropna(subset=['char_treatment', 'drug_ctrl_treatment_key_word', 'gene_cyto_treatment_key_word'], how='all').reset_index(drop=True)  ## del 0 rows
    

    drug_list = pd.read_csv(drug_list_path,sep='\t',header=None)
    drug_list=sorted(drug_list[0].to_list())
    gene_key = pd.read_csv(gene_key_path,sep='\t',header=None)
    gene_key = sorted(gene_key[0].to_list())
    control = pd.read_csv(control_path,sep='\t',header=None)
    control = sorted(control[0].to_list())
    cyto_key = load_workbook(filename=cyto_key_path)
    cyto_key = cyto_key[cyto_key.sheetnames[0]]
    cyto_key_new =[]
    for row in cyto_key.iter_rows(values_only=True):
        print(row)
        cyto_key_new.append(row)

    result_list = remove_parentheses(cyto_key_new)
    cyto_key_new = sorted(result_list)

    treatment_level_dict = create_sorted_dict(drug_list, gene_key, cyto_key_new, control)

    ############################# 上面那种方法是产生括号的那种键值对，需要产另一种
    drug_list = [item.lower() for item in drug_list]
    cyto_key_new = [item.lower() for item in cyto_key_new]
    # control = [item.lower() for item in control]  # control这个不需要变小写，因为有全大写的
    combined_dict = {item: 'drug' for item in drug_list}
    combined_dict.update({item: 'gene' for item in gene_key})
    combined_dict.update({item: 'cyto' for item in cyto_key_new})
    combined_dict.update({item: 'control' for item in control})
    # 排序，首先按值排序，其次对值相同的key排序
    treatment_level_dict_one = dict(sorted(combined_dict.items(), key=lambda x: (x[1], x[0])))

    root = ET.Element("Treatment")
    level1_dict = {}
    for level2, level1 in treatment_level_dict.items():
        if level1 not in level1_dict:
            # 如果level1 tissue不在字典中，创建一个新的level1元素
            level1_element = ET.SubElement(root, "L1", name=level1)
            level1_dict[level1] = level1_element
        # 为level2 tissue创建一个子元素
        ET.SubElement(level1_dict[level1], "L2", name=level2)

    tree = ET.ElementTree(root)
    xml_filename = "treat_level.xml"
    tree.write(xml_filename, encoding='utf-8', xml_declaration=True)
    print(f"XML 文件已生成并保存为 {xml_filename}")


    gse_gsm_clean['treat_char_new'] = gse_gsm_clean['char_treatment'].apply(lambda x: match_treat(x, treatment_level_dict_one))   ### 分别构建多个列，然后相互补全
    ## process the drug & control, gene & cyto columns
    gse_gsm_clean['treat_drug_new'] = gse_gsm_clean['drug_ctrl_treatment_key_word'].apply(lambda x: match_treat(x, treatment_level_dict_one))   ### 分别构建多个列，然后相互补全
    gse_gsm_clean['treat_gene_new'] = gse_gsm_clean['gene_cyto_treatment_key_word'].apply(lambda x: match_treat(x, treatment_level_dict_one))   ### 分别构建多个列，然后相互补全
    gse_gsm_clean['treat_char_key'] = gse_gsm_clean['char_treatment'].apply(lambda x: match_treat_key(x, treatment_level_dict_one))   ### 分别构建多个列，然后相互补全
    ## process the drug & control, gene & cyto columns
    gse_gsm_clean['treat_drug_key'] = gse_gsm_clean['drug_ctrl_treatment_key_word'].apply(lambda x: match_treat_key(x, treatment_level_dict_one))   ### 分别构建多个列，然后相互补全
    gse_gsm_clean['treat_gene_key'] = gse_gsm_clean['gene_cyto_treatment_key_word'].apply(lambda x: match_treat_key(x, treatment_level_dict_one))   ### 分别构建多个列，然后相互补全
    gse_gsm_clean['treat_char_new'].value_counts()  ### 
    gse_gsm_clean['treat_drug_new'].value_counts()
    gse_gsm_clean['treat_gene_new'].value_counts() ### 将一行中同时有几个信息的给标记出来
    ########### Next need to integrate all the 3 items to get the treatment information
    # gse_gsm_clean_new = gse_gsm_clean[['gse','gsm','treat_char_new', 'treat_drug_new', 'treat_gene_new', 'treat_char_key','treat_drug_key', 'treat_gene_key']]
    word_priority_columns = [
        'treat_drug_key', 
        'treat_gene_key', 
        'treat_char_key'
    ]
    type_priority_columns =['treat_drug_new', 
        'treat_gene_new', 
        'treat_char_new']
    
    
    # 应用函数并创建新列treat_final
    gse_gsm_clean['treat_type'] = gse_gsm_clean.apply(get_treat_type, axis=1)  ## 确认过都是的
    gse_gsm_clean['treat_final'] = gse_gsm_clean.apply(get_treat_final, axis=1) 
    gse_gsm_clean['treat_final'].value_counts()   ## 524283(none), 176125(control), 38834(papain), 16537(control;hepes;isoflurane;taurine;trehalose)
    gse_gsm_clean['treat_char_new'].value_counts()  # 950632(none), 20193(control), 11961( cyto), 4409( drug),1051(cyto,drug), 163(control,cyto),154(cyto,gene), 150(gene)
    gse_gsm_clean['treat_drug_new'].value_counts()  ## 548624(none), 194312(drug), 174810(control), 69329(control,drug), 1153(drug,gene), 443(cyto,drug), 255(control;cyto;drug)
    gse_gsm_clean['treat_gene_new'].value_counts()  ## 949619(none), 29382(cyto), 9615(gene), 310(cyto,gene)  
    gse_gsm_clean_new = gse_gsm_clean[['gse','gsm','char_treatment','drug_ctrl_treatment_key_word','gene_cyto_treatment_key_word','treat_type','treat_final']]


    
    gse_ids = sorted(list(gse_gsm_clean_new['gse'].unique()))    # 9246(total)
    # gse_gse_treat = pd.DataFrame(columns=list(gse_gsm_clean_new.columns))
    tmp_list = []
    for i in tqdm(range(len(gse_ids))):  
        treat_tmp = gse_gsm_clean_new[gse_gsm_clean_new['gse']==gse_ids[i]]
        treat_tmp = treat_tmp[~((treat_tmp['gsm'].str.contains('GSE')) & (treat_tmp['treat_final']=='none'))]  ## 经check， treat_final为'none'的,treat_type也为'none'
        if (~treat_tmp['gsm'].str.contains('GSE')).all():       ##    (any(tissue_tmp['gsm'].str.contains('GSE'))), any 是检查至少有一个 GSE
            treat_type = list(treat_tmp['treat_type'].unique())
            treat_final = list(treat_tmp['treat_final'].unique())
            if ((len(treat_type)==1) & len(treat_final)==1):   #### 就是在 tissue_a, 这个coverage，还是看下这个
                # print('yes')
                a = pd.DataFrame(treat_tmp.iloc[0:1])   ## tissue_tmp.iloc[0] iloc 得到的是 Series 对象
                a['gsm'] =gse_ids[i]
                tmp_list.append(a)
            else:
                print(gse_ids[i])

    gse_gse_treat = pd.concat(tmp_list, ignore_index=True)

    
    gse_gse_treat = gse_gse_treat[(gse_gse_treat['treat_final']!='none') | (gse_gse_treat['treat_type']!='none')]    ###  这里只需要去掉 treat_type 和treat_final都为 'none'的行就好，别的可以先保留着
    gse_gsm_clean_new[gse_gsm_clean_new['gsm'].str.contains('GSE')][(gse_gsm_clean_new['treat_final']!='none') & (gse_gsm_clean_new['treat_type']!='none')]  # ｜ 也是 4019 rows

    gse_gsm_treat_rectify  = gse_gsm_clean_new.copy()
    for idx, row in gse_gsm_clean_new.iterrows():
        matched_row = gse_gse_treat[(gse_gse_treat['gse']==row['gse']) & (gse_gse_treat['gsm']==row['gsm']) ]
        if not matched_row.empty:
            gse_gsm_treat_rectify.iloc[idx] = matched_row.iloc[0]

    gse_gsm_treat_rectify[gse_gsm_treat_rectify['gsm'].str.contains('GSE')][(gse_gsm_treat_rectify['treat_final']!='none') & (gse_gsm_treat_rectify['treat_type']!='none')]  # ｜ 也是 5733 rows
    gse_gsm_treat_rectify['treat_final'].value_counts() 
    gse_gsm_treat_rectify['treat_type'].value_counts() 
    
    
    gse_gsm_treat_rectify['treat_final'] = gse_gsm_treat_rectify.groupby('gsm')['treat_final'].transform(fill_none_with_value)
    gse_gsm_treat_rectify['treat_type'] = gse_gsm_treat_rectify.groupby('gsm')['treat_type'].transform(fill_none_with_value)

    gse_gsm_treat_rectify.to_csv(gse_gsm_treatment_output_csv,sep=',',index=False)

    gse_treat_rectify = gse_gsm_treat_rectify[gse_gsm_treat_rectify['gsm'].str.contains('GSE')]

    gse_treat_rectify.to_csv(gse_treatment_output_csv,sep=',',index=False)

    return gse_gsm_treat_rectify, gse_treat_rectify



def human_age_correct(age_meta_path:str,
                      output_path:str):
    data = pd.read_csv(age_meta_path)
    # data.iloc[2,2]
    data = data[data['Species'].apply(lambda x: x == "['Homo sapiens']")]
    data = data.reset_index(drop=True)
    data_path = output_path
    os.makedirs(data_path, exist_ok=True)
    
    with open(os.path.join(data_path,'human_accuracy.csv'), 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['gse', 'gsm', 'Species','char_age_key','char_age','year_old','Label']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader() 
        
    dict_list = []
    conversion_factors = {
        'month': 1/12,   # 1个月约为0.0833年
        '(months)': 1/12,
        'months':1/12,
        'year': 1,    '(years)': 1,   'years':1,  'yr':1,'yrs':1,# 1年
        'day': 1/365  ,  '(days)': 1/365  , 'days':1/356,# 1天约为0.0027年
        'week':1/52,'(weeks)':1/52,'weeks':1/52,
        'hours':1/24/365,
        'w':1/52,
        'd':1/365  ,
        'y':1,
        'm':1/12,
        'h':1/24/365
    }
    conversion_keys_set = set(['month','months','years','year','day','days','week','weeks','w','d','y','p','m',
                           '(months)','(weeks)','(years)','(days)','hours','yr','yrs'])

    fully_conversion_factors = {
        'month': 1/12,   # 1个月约为0.0833年
        '(months)': 1/12,'months':1/12,
        'year': 1,    '(years)': 1,    'years':1, 'yr':1, 'yrs':1,# 1年,
        'day': 1/365  ,  '(days)': 1/365  , 'days':1/365,# 1天约为0.0027年
        'week':1/52,'(weeks)':1/52, 'weeks':1/52,
        'hours':1/24/365}

    Age_key={'fetal':'A','fertilization':"A",'gestational':"A",'cord blood':'A','embryo':'A',
            'spermatid':'A','somite':'A','yolk sac':'A','blastocyst':"A",
            'pediatric':"A",'infant':'A','adult':'Adult','gastrulation':'A','spermatocyto':'A',
            "child":'A','embrionic':'A','juvenile':"A",'pregnancy':'A','puberty':"A",
            'adolescent':"A",'episc':'A','esc':'A',"epilc":'A',"ipscs":'A',
            'neonatal':'A','spermatocytes':'A','unfertilized oocyte':'A',"Zygote":'A',
            'young':'B','frail':"F",'neonate':'A','newborn':'A','pcw':'A',
            'not provided':'Unknown','unknown':'Unknown','not specified':'Unknown','unspecified':'Unknown','mix':"Unknown",
            '2-cell':'A','4-cell':'A','8-cell':'A','16-cell':'A','2 cell':'A','4 cell':'A',
            '8 cell':'A','16 cell':'A','morula':'A','toddler':'A','oocyte':'A','fetus':'A',
            'endoderm':'A','blastomere':'A','ectoderm':'A','gestation':'Adult','lactation':'Adult',
            'trophoblast':'A','spermatogonia':'A','postnal':'Adult','postnatal':'Adult','blastula':'A',
            'mesoderm':'A','germinal vesicle':'A','?':'Unknown','older':'F','gestation':'A'}


    complete_Age_key = {'aged':'F','old':"F",'na':'Unknown','birth':"A"}

    word_to_number = {
        "zero": "0",
        "one": "1",
        "two": "2",
        "three": "3",
        "four": "4",
        "five": "5",
        "six": "6",
        "seven": "7",
        "eight": "8",
        "nine": "9",
        "ten": "10",
        "eleven": "11",
        "twelve": "12",
        "thirteen": "13",
        "fourteen": "14",
        "fifteen": "15",
        "sixteen": "16",
        "seventeen": "17",
        "eighteen": "18",
        "nineteen": "19",
        "twenty": "20"
    }

    def age_to_stage(age):
        if 0<=age<20:
            return 'A'
        elif 20 <= age < 30:
            return 'B'
        elif 30 <= age < 40:
            return 'C'
        elif 40 <= age < 50:
            return 'D'
        elif 50 <= age < 60:
            return 'E'
        elif 60 <= age < 70:
            return 'F'
        elif 70 <= age < 90:
            return 'G'
        elif age >= 90:
            return 'H'
        
    def median_of_range(range_str):
        # 使用正则表达式提取字符串中的数字
        numbers = re.findall(r'\d+', range_str)
        if len(numbers) != 2:
            raise ValueError("Input string should contain exactly two numbers separated by a dash.")
        # 将字符串转换为整数
        num1 = int(numbers[0])
        num2 = int(numbers[1])
        # 计算中位数
        median = (num1 + num2) / 2
        return median

    def average_age(string):
        # 使用正则表达式找到匹配的数字区间
        match = re.search(r'(\d+)-(\d+)', string)
        if match:
            start = int(match.group(1))  # 第一个数字
            end = int(match.group(2))    # 第二个数字
            average = (start + end) / 2  # 计算平均数
            average_str = f'{average:.1f}'  # 将平均数格式化为一位小数的字符串
            return string.replace(f'{start}-{end}', average_str)  # 返回替换后的字符串
        else:
            return string 
        
    for i in range(data.shape[0]):
        gse = data.loc[i,'gse']
        gsm = data.loc[i,'gsm']
        Species = data.loc[i,'Species']
        char_age_key = data.loc[i,'char_age_key']
        char_age = data.loc[i,'char_age']
        age_old = ''
        button = ''
        try:
        ##Check for the presence of NA
            if pd.isna(data.loc[i, 'char_age']):
                    age_old = 'Unknown'
                    with open(os.path.join(data_path,'human_accuracy.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age,'year_old':'Unknown','Label':age_old})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                        print(f"正确，将第 {i} 行数据写入 human_accuracy.csv")
                        continue
            
            ##英文数字变成数字(1-20)
            for word in word_to_number.keys():
                if word in char_age:
                    char_age = char_age.replace(word,word_to_number[word])
            
            ##是否存在‘na、old、aged字符’
            for complete_age_key in complete_Age_key.keys():
                #if re.search(f'\\b{complete_age_key}\\b', char_age.lower()):
                if char_age.lower() in complete_age_key:
                    age_old = complete_Age_key[complete_age_key]
                    with open(os.path.join(data_path,'human_accuracy.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age,'year_old':age_old,'Label':age_old})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                        print(f"正确，将第 {i} 行数据写入 human_accuracy.csv")
                    break
            
            if age_old != '':
                    continue
        
            ##是否存在常见字符
            for age_key in Age_key.keys():
                if age_key in char_age.lower() or age_key in char_age_key.lower():
                    print(f'{age_key}')
                    age_old = Age_key[age_key]
                    with open(os.path.join(data_path,'human_accuracy.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age,'year_old':age_old,'Label':age_old})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 human_accuracy.csv")
                    break

            if age_old != '':
                    continue
            
            if 'passage' in char_age_key or 'passage' in char_age or 'time' in char_age_key or 'time' in char_age  :
                with open(os.path.join(data_path,'human_accuracy.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age,'year_old':'Unknown','Label':'Unknown'})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                        print(f"正确，将第 {i} 行数据写入 human_accuracy.csv")
                with open(os.path.join(data_path,'human_cell_line.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                        print(f"cell_line,将第 {i} 行数据写入 human_cell_line.csv")
                continue
            
            keys = char_age_key.split(',')
            values = char_age.split(',')
            if len(keys) >= len(values):
                result = {keys[j].strip(): values[j].strip() for j in range(len(keys))}
            else:
                result = {keys[j].strip(): values[j].strip() for j in range(len(values))}
            for word_key, word_value in result.items():
                #word_key_set = set(word_key.split())
                word_value_set = set(word_value.split())
                print(word_value_set)
                #word_key_intersection = conversion_keys_set.intersection(word_key_set)
                intersection = conversion_keys_set.intersection(word_value_set)
                #intersection = word_key_intersection.union(word_value_intersection)
                if len(intersection) > 1 :
                    age_old = 0
                    for word in intersection:
                            number_1 = 0
                            match_1 = re.search(rf'(\d+) {word}', word_value)
                            number_1 = int(match_1.group(1)) if match_1 else 0
                            number_1_value = number_1*conversion_factors[word]
                            age_old = number_1_value + age_old
                    continue
                print(f"Key: {word_key}, Value: {word_value}")
                #if 'age' in word_key :
                if '-' in word_value or 'to' in word_value:
                        #age_old = median_of_range(word_value)
                        word_value = average_age(word_value)

                if '+' in word_value:
                    age_old = 0
                    word_value_list = word_value.split('+')
                    for words in word_value_list:
                        for key_word in conversion_factors.keys():
                            if key_word in words.lower():
                                number_1 = 0
                                match = re.search(r'(\d+)', words)
                                if match:
                                    integer_value = float(match.group(0))
                                    number_1 = integer_value * conversion_factors[key_word]
                                    age_old = age_old + number_1
                    break
                if word_value.replace('.', '').isdigit() and len(intersection)<=1 and float(word_value) >= 110:
                    age_old = 0
                    #print(f"{word_value} year")
                    break
                #if '.' in word_value:
                if word_value.replace('.', '').isdigit()and len(intersection)<=1 and float(word_value) < 110:
                    for key_word in fully_conversion_factors.keys():
                        if key_word in word_key.lower():
                            print(f"{key_word}")
                            age_old = float(word_value) * fully_conversion_factors[key_word]
                            break
                        else:
                            age_old = float(word_value) 
                    #print(f"{word_value} year")
                    break

                for key_word in conversion_factors.keys():
                    if key_word in word_value.lower():
                        print(f'{key_word}')
                        if '.' in word_value:
                            match = re.search(r'(\d+\.\d+|\d+)', word_value) 
                            if match:
                                integer_value = float(match.group(0))  # 将匹配到的数字转换为浮点数
                                age_old = float(integer_value) * conversion_factors[key_word]  # 计算年龄
                                break
                        else:
                            match = re.search(r'(\d+)', word_value)
                            if match:
                                integer_value = float(match.group(0))
                                age_old = float(integer_value) * conversion_factors[key_word]
                                #age_old = f"{age_in_years:.3f} year"
                                #print(f"{age_in_years:.3f} year")
                                break
                    if key_word in word_key.lower():
                        print(f"{key_word}")
                        match = re.search(r'(\d+)', word_value)
                        if match:
                            integer_value = float(match.group(0))
                            age_old = float(integer_value) * conversion_factors[key_word]
                            #age_old = f"{age_in_years:.3f} year"
                            #print(f"{age_in_years:.3f} year")
                            break

                if age_old != '':
                    break 

            if age_old != "":
                    final_age_old = age_to_stage(age_old)
                    with open(os.path.join(data_path,'human_accuracy.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age,'year_old':round(age_old,3),'Label':final_age_old})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                        print(f"正确，将第 {i} 行数据写入 human_accuracy.csv")
            else:
                with open(os.path.join(data_path,'human_accuracy.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                    'char_age':char_age,'year_old':'','Label':'Unknown'})
                    #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 human_accuracy.csv")
                with open(os.path.join(data_path,'human_error.csv'), 'a') as f:
                    f.write(','.join(map(str, data.loc[i,:].values)) + '\n')
                    print(f"出错，将第 {i} 行数据写入 human_error.csv:")     
        except Exception as e:
            # 发生异常时将当前行的数据写入 error.csv
            with open(os.path.join(data_path,'human_error.csv'), 'a') as f:
                f.write(','.join(map(str, data.loc[i,:].values)) + '\n')
            print(f"出错，将第 {i} 行数据写入 human_error.csv: {str(e)}")     
            with open(os.path.join(data_path,'human_accuracy.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age,'year_old':'','Label':'Unknown'})
                    #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 human_accuracy.csv") 

  
def mouse_age_correct(
    age_meta_path:str,
    output_path:str):
    data = pd.read_csv(age_meta_path)
    data = data[data['Species'].apply(lambda x: x == "['Mus musculus']")]
    data = data.reset_index(drop=True)
    data
    data_path = output_path
    
    with open(os.path.join(data_path,"mouse_accuracy.csv"), 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['gse', 'gsm', 'Species','char_age_key','char_age','week_old','Label']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader() 
    
    dict_list = []
    conversion_factors = {
        'month': 4.3, '(months)': 4.3,'months':4.3,'(m)':4.3,
        'years':52,'year': 52,'(years)': 52, 'yrs':52,'yr':52, # 1年
        'day': 1/7 , '(days)': 1/7  ,'days':1/7,  # 1天约为0.0027年
        'week':1,'(weeks)':1,'weeks':1,
        'min':1/10080,
        'hour':1/168,"hours":1/168,"h":1/168,
        'p':1/7,
        'd':1/7 ,
        'w':1,
        'm':4.3,
        'y':52, 
    }
    
    conversion_keys_set = set(['month','months','years','year','day','days','week','weeks','w','d','y','p','m',
                           '(months)','(weeks)','(years)','(days)','(m)','yrs','yr'])
    fully_conversion_factors = {
        'month': 4.3, '(months)': 4.3,'months':4.3,'(m)':4.3,
        'years':52,'year': 52,'(years)': 52,    # 1年
        'day': 1/7 , '(days)': 1/7  ,'days':1/7,  # 1天约为0.0027年
        'week':1,'(weeks)':1,'weeks':1,
        'min':1/10080,
        'hour':1/168,"hours":1/168}

    Age_key={'fetal':'Embryo','fertilization':"Embryo",'gestational':"Embryo",'embryo':'Embryo',
            'spermatid':'Embryo','somite':'Embryo','yolk sac':'Embryo','blastocyst':"Embryo",
            'pediatric':"Infant",'infant':'Infant','adult':'Adult','gastrulation':'Embryo','spermatocyto':'Embryo',
            "child":'Children','embrionic':'Embryo','juvenile':"Adolescent",'pregnancy':'Embryo','puberty':"Adolescent",
            'adolescent':"Adolescent",'episc':'Embryo','esc':'Embryo',"epilc":'Embryo',"ipscs":'Embryo',
            'neonatal':'Embryo','spermatocytes':'Embryo','unfertilized oocyte':'Embryo',"zygote":'Embryo',
            'young':'Young Adult','oocyte':"Embryo",'middle':"Adult",'blank':'Unknown','newborn':'Infant',
            'pcw':'Embryo','not provided':'Unknown','unknown':'Unknown','not specified':'Unknown',
            'unspecified':'Unknown','mix':"Unknown",'?':'Unknown','neonate':'Infant', 'gestation':'Adult',
            'lactation':'Adult','spermatogonia':'Embryo','morula':'Embryo','postnal':'Adult','postnatal':'Adult',
            'blastula':'Embryo','mesoderm':'Embryo','ectoderm':'Embryo','endoderm':'Embryo','germinal vesicle':'Embryo',
            'allantoic bud':'Embryo','2-cell':'Embryo','4-cell':'Embryo','8-cell':'Embryo','16-cell':'Embryo',
            '2 cell':'Embryo','4 cell':'Embryo','8 cell':'Embryo','16 cell':'Embryo','trophoblast':'Embryo'}

    complete_Age_key = {'old':"Aged",'aged':'Aged','na':'Unknown','birth':"Infant"}

    word_to_number = {
        "zero": "0",
        "one": "1",
        "two": "2",
        "three": "3",
        "four": "4",
        "five": "5",
        "six": "6",
        "seven": "7",
        "eight": "8",
        "nine": "9",
        "ten": "10",
        "eleven": "11",
        "twelve": "12",
        "thirteen": "13",
        "fourteen": "14",
        "fifteen": "15",
        "sixteen": "16",
        "seventeen": "17",
        "eighteen": "18",
        "nineteen": "19",
        "twenty": "20"
    }

    def mouse_age_to_stage(age):
        if age ==0:
            return 'Embryo'
        if age > 0 and age <= 3:
            return "Infant"
        elif age > 3 and age <= 4:
            return "Children"
        elif age > 4 and age <= 6:
            return "Adolescent"
        elif age > 6 and age <= 14:
            return "Young adult"
        elif age > 14 and age <= 72:
            return "Adult"
        elif age > 72 and age <= 120:
            return "Aged"
        elif age > 120 :
            return "Super old"
        
    def median_of_range(range_str):
        # 使用正则表达式提取字符串中的数字
        numbers = re.findall(r'\d+', range_str)
        if len(numbers) != 2:
            raise ValueError("Input string should contain exactly two numbers separated by a dash.")
        # 将字符串转换为整数
        num1 = int(numbers[0])
        num2 = int(numbers[1])
        # 计算中位数
        median = (num1 + num2) / 2
        return median

    def average_age(string):
        # 使用正则表达式找到匹配的数字区间
        match = re.search(r'(\d+)-(\d+)', string)
        if match:
            start = int(match.group(1))  # 第一个数字
            end = int(match.group(2))    # 第二个数字
            average = (start + end) / 2  # 计算平均数
            average_str = f'{average:.1f}'  # 将平均数格式化为一位小数的字符串
            return string.replace(f'{start}-{end}', average_str)  # 返回替换后的字符串
        else:
            return string 
        
    def bo_average_age(string):
        # 使用正则表达式找到匹配的数字区间
        match = re.search(r'(\d+)~(\d+)', string)
        if match:
            start = int(match.group(1))  # 第一个数字
            end = int(match.group(2))    # 第二个数字
            average = (start + end) / 2  # 计算平均数
            average_str = f'{average:.1f}'  # 将平均数格式化为一位小数的字符串
            return string.replace(f'{start}-{end}', average_str)  # 返回替换后的字符串
        else:
            return string 
  
        
    for i in range(data.shape[0]):
        gse = data.loc[i,'gse']
        gsm = data.loc[i,'gsm']
        Species = data.loc[i,'Species']
        char_age_key = data.loc[i,'char_age_key']
        char_age = data.loc[i,'char_age']
        age_old = ''
        button = ''
        try:
            ##是否存在空值
            if pd.isna(data.loc[i, 'char_age']):
                age_old = 'Unknown'
                with open(os.path.join(data_path,"mouse_accuracy.csv"), 'a',newline='',encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                    'char_age':char_age,'week_old':'Unknown','Label':age_old})
                    #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 mouse_accuracy.csv")
                    continue
            #char_age_key = data.loc[i,'char_age_key']
            #char_age = data.loc[i,'char_age']
            ##英文数字变成数字
            for word in word_to_number.keys():
                if word in char_age.lower():
                    print(f"{word}")
                    char_age = char_age.lower().replace(word,word_to_number[word])
            ##是否有‘Exx’存在
            if re.search(r'e\d+', char_age.lower()):
                match = re.search(r'e\d+', char_age.lower())
                E_value = match.group(0).replace("e",'E')
                age_old = 'Embryo'
                with open(os.path.join(data_path,"mouse_accuracy.csv"), 'a',newline='',encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                    'char_age':char_age,'week_old':E_value,'Label':age_old})
                    #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 mouse_accuracy.csv")
                continue
            ##是否存在‘na、old、aged字符’
            for complete_age_key in complete_Age_key.keys():
                if re.search(r'(?<!\S){}(?!\S)'.format(re.escape(complete_age_key)), char_age.lower()):
                    print(complete_age_key)
                # if char_age.lower() in complete_age_key:
                    age_old = complete_Age_key[complete_age_key]
                    print(age_old)
                    with open(os.path.join(data_path,"mouse_accuracy.csv"), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age,'week_old':age_old,'Label':age_old})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                        print(f"正确，将第 {i} 行数据写入 mouse_accuracy.csv")
                    break
            if age_old != '':
                continue
            ##是否存在常见字符
            for age_key in Age_key.keys():
                if age_key in char_age.lower() or age_key in char_age_key.lower():
                    print(f'{age_key}')
                    age_old = Age_key[age_key]
                    with open(os.path.join(data_path,"mouse_accuracy.csv"), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age,'week_old':age_old,'Label':age_old})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 mouse_accuracy.csv")
                    break
            if age_old != '':
                continue
                
            
            # if ('passage' in char_age_key or 'passage' in char_age) and (not re.search(r'\bage\b', char_age_key) and not re.search(r'\bage\b', char_age))or :
            if ('passage' in char_age_key or 'passage' in char_age) or ('time' in char_age_key or 'time' in char_age):
                with open(os.path.join(data_path,'mouse_cell_line.csv'), 'a',newline='',encoding='utf-8') as csvfile:
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                        'char_age':char_age})
                        #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                        print(f"cell_line,将第 {i} 行数据写入 mouse_cell_line.csv")
                with open(os.path.join(data_path,"mouse_accuracy.csv"), 'a',newline='',encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                    'char_age':char_age,'week_old':'','Label':'Unknown'})
                    #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 mouse_accuracy.csv")
                with open(os.path.join(data_path,"mouse_error.csv"), 'a') as f:
                    f.write(','.join(map(str, data.loc[i,:].values)) + '\n')
                    print(f"出错，将第 {i} 行数据写入 mouse_error.csv:")    
                continue
            
            keys = char_age_key.split(',')
            values = char_age.split(',')
            # if len(keys) >= len(values):
            #     result = {keys[j].strip(): values[j].strip() for j in range(len(values))}
            # else:
            #     result = {keys[j].strip(): values[j].strip() for j in range(len(keys))}
            if len(keys) != len(values):
                result = {char_age_key: char_age}
            else:
                result = {keys[j].strip(): values[j].strip() for j in range(len(values))}
            for word_key, word_value in result.items():
                #word_key_set = set(word_key.split())
                word_value_set = set(word_value.split())
                #word_key_intersection = conversion_keys_set.intersection(word_key_set)
                intersection = conversion_keys_set.intersection(word_value_set)
                #intersection = word_key_intersection.union(word_value_intersection)
                if len(intersection) > 1 :
                    age_old = 0
                    for word in intersection:
                        number_1 = 0
                        match_1 = re.search(rf'(\d+) {word}', word_value)
                        number_1 = int(match_1.group(1)) if match_1 else 0
                        number_1_value = number_1*conversion_factors[word]
                        age_old = number_1_value + age_old
                    continue
                print(f"Key: {word_key}, Value: {word_value}")
                #if 'age' in word_key :
                #if '-' in word_value and not intersection and word_value.replace('-', '').isdigit():
                if '-' in word_value:
                        #age_old = median_of_range(word_value)
                        word_value = average_age(word_value)
                if '+' in word_value:
                    age_old = 0
                    word_value_list = word_value.split('+')
                    for words in word_value_list:
                        for key_word in conversion_factors.keys():
                            if key_word in words.lower():
                                number_1 = 0
                                match = re.search(r'(\d+)', words)
                                if match:
                                    integer_value = float(match.group(0))
                                    number_1 = integer_value * conversion_factors[key_word]
                                    age_old = age_old + number_1
                    break
                #if word_value.isdigit()and not intersection and int(word_value) < 110:
                #                age_old = float(word_value)
                                #print(f"{word_value} year")
                #                break
                if word_value.replace('.', '').isdigit() and len(intersection)<=1 and float(word_value) >= 110:
                                age_old = 0
                                #print(f"{word_value} year")
                                break
                #if '.' in word_value:
                if word_value.replace('.', '').isdigit()and len(intersection)<=1 and float(word_value) < 110:
                                for key_word in fully_conversion_factors.keys():
                                    if key_word in word_key.lower():
                                        print(f"{key_word}")
                                        age_old = float(word_value) * fully_conversion_factors[key_word]
                                        break
                                    else:
                                        age_old = float(word_value) 
                                #print(f"{word_value} year")
                                break
                for key_word in conversion_factors.keys():
                            if key_word in word_value.lower():
                                print(f'{key_word}')
                                if '.' in word_value:
                                    match = re.search(r'(\d+\.\d+|\d+)', word_value) 
                                    if match:
                                        integer_value = float(match.group(0))  # 将匹配到的数字转换为浮点数
                                        age_old = float(integer_value) * conversion_factors[key_word]  # 计算年龄
                                        break
                                else:
                                    match = re.search(r'(\d+)', word_value)
                                    if match:
                                        integer_value = float(match.group(0))
                                        age_old = float(integer_value) * conversion_factors[key_word]
                                        #age_old = f"{age_in_years:.3f} year"
                                        #print(f"{age_in_years:.3f} year")
                                    break
                            if key_word in word_key.lower():
                                print(f"{key_word}")
                                match = re.search(r'(\d+)', word_value)
                                if match:
                                    integer_value = float(match.group(0))
                                    age_old = float(integer_value) * conversion_factors[key_word]
                                    #age_old = f"{age_in_years:.3f} year"
                                    #print(f"{age_in_years:.3f} year")
                                break
                if age_old != '':
                    break       
            if age_old != "":
                # final_age_old = mouse_age_to_stage(age_old)
                final_age_old = str(age_old)+ ' weeks'if age_old != '' else age_old
                with open(os.path.join(data_path,"mouse_accuracy.csv"), 'a',newline='',encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                    'char_age':char_age,'week_old':round(age_old,2),'Label':final_age_old})
                    #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 mouse_accuracy.csv")
            else:
                with open(os.path.join(data_path,"mouse_accuracy.csv"), 'a',newline='',encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                    'char_age':char_age,'week_old':'','Label':'Unknown'})
                    #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 mouse_accuracy.csv")
                with open(os.path.join(data_path,"mouse_error.csv"), 'a') as f:
                    f.write(','.join(map(str, data.loc[i,:].values)) + '\n')
                    print(f"出错，将第 {i} 行数据写入 mouse_error.csv:")     
        except Exception as e:
            # 发生异常时将当前行的数据写入 error.csv
            with open(os.path.join(data_path,"mouse_error.csv"), 'a') as f:
                f.write(','.join(map(str, data.loc[i,:].values)) + '\n')
            print(f"出错，将第 {i} 行数据写入 mouse_error.csv: {str(e)}")   
            with open(os.path.join(data_path,"mouse_accuracy.csv"), 'a',newline='',encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writerow({'gse':gse,'gsm':gsm,'Species':Species,"char_age_key":char_age_key,
                                    'char_age':char_age,'week_old':'','Label':'Unknown'})
                    #f.write(','.join(map(str, data.loc[i,:].values)) + ',' + str(age_old) + '\n')
                    print(f"正确，将第 {i} 行数据写入 mouse_accuracy.csv")  

    
    
    

# ## treatment
# def rectify_treatment_meta(
#         in_csv: str,
#         word_bank_dir: str ,
#         treat_gse_gsm_out: str ,
#         treat_gse_out: str 
# ) -> tuple[pd.DataFrame, pd.DataFrame]:
    
#     # ------------------------------------------------------------------
#     # 1. 读取原始表 & 词表
#     # ------------------------------------------------------------------
#     df = pd.read_csv(in_csv, sep=',')
#     # 只要三列中至少 2 个非空即可保留
#     df = df.dropna(
#         subset=['char_treatment', 'drug_ctrl_treatment_key_word', 'gene_cyto_treatment_key_word'],
#         thresh=2
#     ).reset_index(drop=True)

#     # 读取词表
#     drug_list = sorted(pd.read_csv(os.path.join(word_bank_dir, 'drug.txt'),
#                                    sep='\t', header=None)[0].str.lower().tolist())
#     gene_list = sorted(pd.read_csv(os.path.join(word_bank_dir, 'gene.txt'),
#                                    sep='\t', header=None)[0].str.lower().tolist())
#     cyto_wb = load_workbook(os.path.join(word_bank_dir, 'cytokine.xlsx'))
#     cyto_list = sorted([row[0].lower() for row in cyto_wb[cyto_wb.sheetnames[0]].iter_rows(values_only=True)])
#     control_list = sorted(pd.read_csv(os.path.join(word_bank_dir, 'ctrl.txt'),
#                                       sep='\t', header=None)[0].tolist())

#     ##构建进度条
#     progress_bar = tqdm(total=100, desc='Processing')

#     # 构建映射字典
#     combined_dict = {**{k: 'drug' for k in drug_list},
#                      **{k: 'gene' for k in gene_list},
#                      **{k: 'cyto' for k in cyto_list},
#                      **{k: 'control' for k in control_list}}

#     # ------------------------------------------------------------------
#     # 2. 工具函数
#     # ------------------------------------------------------------------
#     def match_treat(treat_str: str) -> str:
#         if pd.isna(treat_str):
#             treat_str = ''
#         else:
#             treat_str = str(treat_str).lower()
#         match = {v for k, v in combined_dict.items() if k in treat_str}
#         return ';'.join(sorted(match)) if match else 'none'

#     def match_treat_key(treat_str: str) -> str:
#         if pd.isna(treat_str):
#             treat_str = ''
#         else:
#             treat_str = str(treat_str).lower()
#         match = []
#         for k, v in combined_dict.items():
#             if k in treat_str or v in treat_str:
#                 match.append(v if v == 'control' else k)
#         return ';'.join(sorted(set(match))) if match else 'none'

#     # ------------------------------------------------------------------
#     # 3. 逐列提取并整合
#     # ------------------------------------------------------------------
#     df['treat_char_new'] = df['char_treatment'].apply(match_treat)
#     df['treat_drug_new'] = df['drug_ctrl_treatment_key_word'].apply(match_treat)
#     df['treat_gene_new'] = df['gene_cyto_treatment_key_word'].apply(match_treat)
#     progress_bar.update(20)

#     df['treat_char_key'] = df['char_treatment'].apply(match_treat_key)
#     df['treat_drug_key'] = df['drug_ctrl_treatment_key_word'].apply(match_treat_key)
#     df['treat_gene_key'] = df['gene_cyto_treatment_key_word'].apply(match_treat_key)
#     progress_bar.update(40)
#     # 按优先级整合
#     def get_treat_type(row):
#         for col in ['treat_drug_new', 'treat_gene_new', 'treat_char_new']:
#             if row[col] != 'none':
#                 return row[col]
#         return 'none'

#     def get_treat_final(row):
#         for col in ['treat_drug_key', 'treat_gene_key', 'treat_char_key']:
#             if row[col] != 'none':
#                 return row[col]
#         return 'none'

#     df['treat_type'] = df.apply(get_treat_type, axis=1)
#     df['treat_final'] = df.apply(get_treat_final, axis=1)
#     progress_bar.update(60)
#     # ------------------------------------------------------------------
#     # 4. GSE 粒度统一
#     # ------------------------------------------------------------------
#     def unify_per_gse(group_df):
#     # group_df 是当前 GSE 的全部行
#         gse = group_df.name  # 这是分组键
#         mask = ~((group_df['gsm'].str.contains('GSE')) & (group_df['treat_final'] == 'none'))
#         group_df = group_df[mask]

#         if (~group_df['gsm'].str.contains('GSE')).all():
#             type_vals = group_df['treat_type'].unique()
#             final_vals = group_df['treat_final'].unique()
#             if len(type_vals) == 1 and len(final_vals) == 1:
#                 rec = group_df.iloc[:1].copy()
#                 rec['gsm'] = gse
#                 return rec
#         return pd.DataFrame()

#     gse_df = df.groupby('gse', group_keys=False).apply(unify_per_gse)
#     gse_df = gse_df[(gse_df['treat_final'] != 'none') | (gse_df['treat_type'] != 'none')]
#     progress_bar.update(80)
#     # 用 merge/update 方式回填，避免 iloc 越界
#     df = df.merge(
#         gse_df.set_index(['gse', 'gsm'])[['treat_final', 'treat_type']],
#         left_on=['gse', 'gsm'],
#         right_index=True,
#         how='left',
#         suffixes=('', '_gse')
#     )
#     df['treat_final'] = df['treat_final_gse'].fillna(df['treat_final'])
#     df['treat_type'] = df['treat_type_gse'].fillna(df['treat_type'])
#     df = df.drop(columns=['treat_final_gse', 'treat_type_gse'])
#     progress_bar.update(90)
#     # ------------------------------------------------------------------
#     # 5. 整理列并输出
#     # ------------------------------------------------------------------
#     df = df[['gse', 'gsm', 'char_treatment', 'drug_ctrl_treatment_key_word',
#              'gene_cyto_treatment_key_word', 'treat_type', 'treat_final']]
#     df = df.drop_duplicates()

#     df.to_csv(treat_gse_gsm_out, sep=',', index=False)

#     gse_only = df[df['gsm'].str.contains('GSE')]
#     gse_only.to_csv(treat_gse_out, sep=',', index=False)
#     progress_bar.close()
#     return df, gse_only

# ## disease

# # ------------------ 1. 词典 ------------------
# def build_disease_dict(disease_type_csv: str,
#                        disease_name_tsv: str) -> tuple[dict, pd.DataFrame]:
#     """
#     读取 DiseaseType.csv 和 DiseaseType_dictionary_1-multi.tsv
#     返回 (disease_type_dict, disease_type_df)
#     """
#     disease_type = pd.read_csv(disease_type_csv, header=None, names=['Type', 'Label'])
#     disease_type['Label'] = disease_type['Label'].str.replace('\xa0', ' ')
#     disease_type['Type_l1'] = 'Anatomical'
#     disease_type.loc[
#         disease_type['Label'].isin(
#             ['Cancer', 'Fetal disease', 'Genetic disease', 'Infectious disease',
#              'Metabolic disease', 'Rare disease', 'Normal']),
#         'Type_l1'] = 'Global'

#     disease_name = pd.read_csv(disease_name_tsv, sep='\t')
#     disease_dict = {}
#     for _, r in disease_name.iterrows():
#         word = r['Disease']
#         if not (word.isupper() or (len(word) > 2 and word[1:-1].isupper())):
#             word = word.lower()
#         disease_dict[word] = str(r['Type'])
#     disease_dict = {k.lower(): v for k, v in disease_dict.items()}
#     return disease_dict, disease_type


# # ------------------ 2. 疾病标准化 ------------------
# def annotate_disease(raw_df: pd.DataFrame,
#                      disease_dict: dict,
#                      disease_type_df: pd.DataFrame) -> pd.DataFrame:
#     """
#     输入：原始 gse_gsm_disease 表（至少含 gse, gsm, char_disease, disease_key_word）
#     输出：带 Global_type / Anatomical_type 的新表
#     """
#     # 0) 复制避免链式赋值
#     df = raw_df.copy()

#     # 1) 小写处理
#     for col in ['char_disease', 'disease_key_word']:
#         df[col] = df[col].astype(str).str.lower()

#     # 2) 关键词列表
#     normal_kw = ['normal', 'healthy', 'disease-free', 'disease free', 'healty',
#                  'control', 'uninfected', 'no-covid', 'non covid', 'health', 'uninjured']
#     tumor_kw = ['benign', 'metastasis', 'carcinoma', 'tumor', 'cancer',
#                 'adenocarcinoma', 'leiomyoma', 'liposarcoma', 'sarcoma', 'adenoma']

#     # 3) 从 char_disease 提取 type
#     def _char_type(row):
#         cd = str(row['char_disease']).lower()
#         t = [disease_dict[cd]] if cd in disease_dict else []
#         return ','.join(sorted(set(t))) if t else 'none'
#     df['char_type'] = df.apply(_char_type, axis=1)

#     # 4) 从 disease_key_word 提取 type
#     def _disease_type(row):
#         try:
#             kw = ast.literal_eval(row['disease_key_word'])
#             kw = list({str(k).lower() for k in kw})
#         except Exception:
#             kw = []
#         t = [disease_dict[k] for k in kw if k in disease_dict]
#         return ','.join(sorted(set(t))) if t else 'none'
#     df['disease_type'] = df.apply(_disease_type, axis=1)

#     # 5) 关键词兜底
#     def _kw_fallback(text):
#         if pd.isna(text):
#             return 'none'
#         text = str(text).lower()
#         if any(n in text for n in normal_kw):
#             return '25'
#         if any(t in text for t in tumor_kw):
#             return '19'
#         return 'none'

#     df['char_disease_new'] = df['char_disease'].apply(_kw_fallback)

#     # 6) 最终 type
#     def _final(row):
#         if row['disease_type'] != 'none':
#             return row['disease_type']
#         if row['char_type'] != 'none':
#             return row['char_type']
#         if row['char_disease_new'] != 'none':
#             return row['char_disease_new']
#         return 'none'
#     df['disease_final'] = df.apply(_final, axis=1)

#     # 7) 把 type 转成 label
#     type2label = disease_type_df.set_index('Type')['Label'].to_dict()
#     type2label.update({'none': 'None'})
#     def _map(types):
#         labels = [type2label[t] for t in types.split(',') if t in type2label]
#         return ','.join(sorted(set(labels))) if labels else 'None'
#     df['disease_name'] = df['disease_final'].apply(_map)

#     # 8) Global / Anatomical 拆分
#     global_labs = set(disease_type_df[disease_type_df['Type_l1'] == 'Global']['Label'])
#     anatomical_labs = set(disease_type_df[disease_type_df['Type_l1'] == 'Anatomical']['Label'])

#     def _split_global(names):
#         if names == 'None':
#             return 'None'
#         labs = set(names.split(','))
#         g = sorted(labs & global_labs)
#         return ','.join(g) if g else 'None'

#     def _split_anatomical(names):
#         if names == 'None':
#             return 'None'
#         labs = set(names.split(','))
#         a = sorted(labs & anatomical_labs)
#         return ','.join(a) if a else 'None'

#     df['Global_type'] = df['disease_name'].apply(_split_global)
#     df['Anatomical_type'] = df['disease_name'].apply(_split_anatomical)

#     # 9) 整理列
#     out = df[['gse', 'gsm', 'char_disease', 'disease_key_word',
#               'Global_type', 'Anatomical_type']].copy()
#     return out


# # ------------------ 3. 同 GSE 回填 ------------------
# def fill_missing_disease(df: pd.DataFrame) -> pd.DataFrame:
#     """
#     对 Global_type / Anatomical_type 为 'None' 的行，
#     用同一 GSE 内已存在的非 None 值回填
#     """
#     df = df.copy()

#     def _best_fill(ser):
#         """优先无逗号、非 None；其次有逗号、非 None"""
#         no_comma = [x for x in ser if x != 'None' and ',' not in str(x)]
#         with_comma = [x for x in ser if x != 'None' and ',' in str(x)]
#         if no_comma:
#             return no_comma[0]
#         if with_comma:
#             return with_comma[0]
#         return 'None'

#     for gse_id, sub in df.groupby('gse'):
#         g_fill = _best_fill(sub['Global_type'].tolist())
#         a_fill = _best_fill(sub['Anatomical_type'].tolist())
#         mask = (df['gse'] == gse_id) & (df['Global_type'] == 'None')
#         df.loc[mask, 'Global_type'] = g_fill
#         mask = (df['gse'] == gse_id) & (df['Anatomical_type'] == 'None')
#         df.loc[mask, 'Anatomical_type'] = a_fill
#     return df


# # ------------------ 4. 一键跑 ------------------
# def rectify_disease_meta(raw_csv: str,
#                         disease_type_csv: str,
#                         disease_name_tsv: str,
#                         disease_gse_gsm_out_csv: str,
#                         disease_gse_out_csv:str):
#     """
#     一键读取、标注、回填、写出
#     """
#     raw = pd.read_csv(raw_csv)
#     ddict, dtype_df = build_disease_dict(disease_type_csv, disease_name_tsv)
#     ann = annotate_disease(raw, ddict, dtype_df)
#     filled = fill_missing_disease(ann)

#     filled.to_csv(disease_gse_gsm_out_csv, index=False)
    
#     gse_df = filled[filled['gsm'].str.contains('GSE')].copy()
#     gse_df.to_csv(disease_gse_out_csv, index=False)
    
    
#     return filled, gse_df


# ## tissue

# def rectify_tissue_meta(input_file, 
#                         simplify_file, 
#                         cancer_cellline_file, 
#                         gse_gsm_output_file, 
#                         gse_output_file):
#     """
#     Process tissue data from the input file and rectify it based on the provided dictionaries and rules.
    
#     Parameters:
#     - input_file: Path to the input CSV file containing tissue metadata.
#     - simplify_file: Path to the Simplify_new.txt file containing tissue type mappings.
#     - cancer_cellline_file: Path to the cancer_cellline.txt file containing cancer cell lines.
#     - output_file: Path to save the rectified tissue metadata CSV file.
#     """
    
#     # Load the input data
#     gse_gsm_tissue = pd.read_csv(input_file, sep=',')
#     gse_gsm_tissue = gse_gsm_tissue.dropna(subset=['tissue_key_word'])
#     gse_gsm_tissue = gse_gsm_tissue[gse_gsm_tissue['tissue_key_word'] != '']
#     gse_gsm_tissue['tissue_key_word'] = gse_gsm_tissue['tissue_key_word'].str.lower()
    
#     # Load the tissue level data
#     tissue_level = pd.read_csv(simplify_file, sep='\t', header=None)
#     tissue_level.columns = ['tissue', 'tissue_type']
#     tissue_level = tissue_level[tissue_level['tissue_type'] != 'carcinoma']
    
#     # Create tissue dictionaries
#     def split_and_dict(df):
#         result = []
#         for index, row in df.iterrows():
#             tissues = row['tissue'].lower().split(',')
#             tissue_dict = {tissue.strip(): row['tissue_type'] for tissue in tissues}
#             result.append(tissue_dict)
#         return result
    
#     tissue_dict_type = split_and_dict(tissue_level)
    
#     def create_dict_from_tissue_df(df):
#         result_dict = {}
#         for index, row in df.iterrows():
#             tissues = row['tissue'].split(',')
#             processed_words = []
#             for word in tissues:
#                 if word.isupper() or (word[1:-1].isupper() and word[-1].islower()):
#                     processed_words.append(word)
#                 else:
#                     processed_words.append(word.lower())
#             unique_tissues = set(sorted(processed_words))
#             for tissue in unique_tissues:
#                 result_dict[tissue] = row['tissue_type']
#         return result_dict
    
#     tissue_dict_one = create_dict_from_tissue_df(tissue_level)
#     unique_tissue_l1 = sorted(list(set(tissue_dict_one.values())))
    
#     def process_list(my_list):
#         target_set = {'tumor', 'cancer', 'h4', 'u2', 'carcinoma', 'del'}
#         filtered_list = [item for item in my_list if item not in target_set]
#         return filtered_list
    
#     def match_tissue(tissue_str, tissue_dict):
#         if not isinstance(tissue_str, str):
#             tissue_str = str(tissue_str).lower()
#         else:
#             tissue_str = tissue_str.lower()
#         match_value = []
#         for key, value in tissue_dict.items():
#             if key in tissue_str or value in tissue_str:
#                 match_value.append(value)
#         match_value = list(sorted(set(match_value)))
#         return ';'.join(match_value) if match_value else 'none'
    
#     def replace_and_join_tissue(keywords_list1):
#         unique_keywords1 = list(set(ast.literal_eval(keywords_list1)))
#         new_list = []
#         for tissue in unique_keywords1:
#             if tissue in tissue_dict_one:
#                 new_list.append(tissue_dict_one[tissue])
#             else:
#                 new_list.append(tissue)
#         unique_tissues = sorted(list(set(new_list)))
#         if len(unique_tissues) == 1:
#             return ';'.join(unique_tissues)
#         elif len(unique_tissues) > 1 and not set(unique_tissues) & set(unique_tissue_l1):
#             result = process_list(unique_tissues)
#             unique_tissues_1 = sorted(result)
#             if len(unique_tissues_1) == 0:
#                 unique_tissues_1 = ['none']
#             return ';'.join(unique_tissues_1)
#         elif len(unique_tissues) > 1 and set(unique_tissues).issubset(set(unique_tissue_l1)):
#             if 'none' in unique_tissues:
#                 unique_tissues.remove('none')
#             return ';'.join(unique_tissues)
#         else:
#             common_elements = [item for item in unique_tissues if item in set(unique_tissue_l1)]
#             common_elements = sorted(common_elements)
#             if len(common_elements) == 1:
#                 return ';'.join(common_elements)
#             if len(common_elements) > 1 and 'none' in common_elements:
#                 common_elements.remove('none')
#             return ';'.join(common_elements)
    
#     gse_gsm_tissue['tissue_new'] = gse_gsm_tissue['tissue_key_word'].apply(replace_and_join_tissue)
#     gse_gsm_tissue['tissue_new_char'] = gse_gsm_tissue['char_tissue'].apply(lambda x: match_tissue(x, tissue_dict_one))
    
#     def replace_none_with_tissue(df):
#         new_df = df.copy()
#         for index, row in new_df.iterrows():
#             if pd.isna(row['tissue_new_char']) or row['tissue_new_char'].lower() == 'none':
#                 new_df.loc[index, 'tissue_new_char'] = row['tissue_new']
#         return new_df
    
#     result_df = replace_none_with_tissue(gse_gsm_tissue)
#     result_df.rename(columns={'tissue_new_char': 'tissue_final'}, inplace=True)
#     result_df['tissue_final'] = result_df['tissue_final'].fillna('none')
#     result_df['tissue_final'] = result_df['tissue_final'].str.replace('bone;bone marrow', 'bone marrow')
    
#     # Load cancer cellline data
#     cancer_cellline = pd.read_csv(cancer_cellline_file, sep='\t', header=None)
#     cancer_cellline.columns = ['Cancer cellline']
    
#     def assign_tissue_type(df, cancer_cellline, unique_tissue_l1):
#         def check_tissue_type(tissue_final_str):
#             tissue_list = [t.strip() for t in tissue_final_str.split(';')]
#             if tissue_list == ['none']:
#                 return 'none'
#             elif any(t in cancer_cellline for t in tissue_list):
#                 return 'cancer cellline'
#             elif any(t in unique_tissue_l1 for t in tissue_list):
#                 return 'tissue'
#             else:
#                 return 'cellline'
#         df['tissue_type'] = df['tissue_final'].apply(check_tissue_type)
#         return df
    
#     gse_gsm_tissue_processed = assign_tissue_type(result_df, cancer_cellline['Cancer cellline'].to_list(), unique_tissue_l1)
#     gse_gsm_tissue_processed.to_csv(gse_gsm_output_file, sep=',', index=False)
    
#     gse_tissue_processed = gse_gsm_tissue_processed[gse_gsm_tissue_processed['gsm'].str.contains('GSE')]
#     gse_tissue_processed.to_csv(gse_output_file, sep=',', index=False)
    
#     return gse_gsm_tissue_processed, gse_tissue_processed


# def rectify_platform_meta(
#         in_csv: str,
#         mapping_tsv: Optional[str] = None,
#         platform_gse_gsm_out_csv: str = 'gse_gsm_platform_meta_rectify_test.csv',
#         platform_gse_out: str = 'gse_platform_meta_rectify_test.csv'
# ) -> Tuple[pd.DataFrame, pd.DataFrame]:
   
#     # ------------------------------------------------------------------
#     # 0. 路径准备
#     # ------------------------------------------------------------------
#     mapping_tsv = mapping_tsv or os.path.join(os.path.dirname(__file__),
#                                               'word_bank', 'platform_mapping_list.tsv')

#     # ------------------------------------------------------------------
#     # 1. 读取并初步清洗
#     # ------------------------------------------------------------------
#     df = pd.read_csv(in_csv, sep=',')
#     df = df.dropna(subset=['platform_key_word']).reset_index(drop=True)
#     df = df[df['platform_key_word'] != '']
#     df['platform_key_word'] = df['platform_key_word'].str.lower()

#     # 映射表
#     mapping = pd.read_csv(mapping_tsv, sep='\t')
#     mapping = mapping.drop(columns=['Remarks']).drop_duplicates()
#     mapping['Platform'] = mapping['Platform'].str.lower()
#     mapping = mapping.drop_duplicates()

#     platform_dict = dict(zip(mapping['Platform'], mapping['Platform_rename']))
#     platform_dict = {k.lower(): v for k, v in platform_dict.items()}
#     unique_platform_l2 = list(set(platform_dict.values()))
#     if 'droplet/well-based' in unique_platform_l2:
#         unique_platform_l2.remove('droplet/well-based')

#     # ------------------------------------------------------------------
#     # 2. 工具函数
#     # ------------------------------------------------------------------
#     def process_smart(lst):
#         targets = {'Smart-seq', 'Smart-seq2', 'SMARTer'}
#         return sorted(set([i for i in lst if i not in targets] + ['Smart-seq2']))

#     def replace_and_join(keywords_str):
#         kw = set(ast.literal_eval(keywords_str))
#         kw = {k.lower() for k in kw}
#         new = [platform_dict.get(k, k) for k in kw]
#         new = [n for n in new if n != 'droplet/well-based']

#         if not new:
#             return 'none'
#         if len(new) == 1:
#             return new[0]

#         if '10X Genomics' in new:
#             return '10X Genomics'
#         if 'CITE-seq' in new:
#             new = [i for i in new if i != 'CITE-seq']
#         new = process_smart(new)
#         return ';'.join(sorted(set(new)))

#     def rectify_char_platform(x):
#         if pd.isna(x):
#             return 'none'
#         x = str(x).lower()
#         if '10x' in x:
#             return '10X Genomics'
#         if 'smart' in x:
#             return 'Smart-seq2'
#         return 'none'

#     # ------------------------------------------------------------------
#     # 3. 生成 platform_new（key_word 列）
#     # ------------------------------------------------------------------
#     df['platform_new'] = df['platform_key_word'].apply(replace_and_join)

#     # 4. 基于 char_platform 回填
#     df['platform_new_char'] = df['char_platform'].apply(rectify_char_platform)
#     mask = df['platform_new'].isin(['none', '', np.nan])
#     df.loc[mask, 'platform_new'] = df.loc[mask, 'platform_new_char']

#     # ------------------------------------------------------------------
#     # 5. 按 GSE 粒度统一
#     # ------------------------------------------------------------------
#     def unify_per_gse(group):
#         platforms = [p for p in group['platform_new'].unique() if p not in {'none', 'droplet/well-based'}]
#         if not platforms:
#             unified = 'none'
#         elif '10X Genomics' in platforms:
#             unified = '10X Genomics'
#         elif any('Smart-seq' in p for p in platforms):
#             unified = 'Smart-seq2'
#         else:
#             unified = platforms[0]

#         group.loc[group['gsm'].str.contains('GSE'), 'platform_new'] = unified
#         return group

#     df = df.groupby('gse', group_keys=False).apply(unify_per_gse)

#     # ------------------------------------------------------------------
#     # 6. 后处理：替换、去 none_list、重命名
#     # ------------------------------------------------------------------
#     replacements = {
#         'Drop-seq;droplet/well-based': 'Drop-seq',
#         'SureCell;droplet/well-based': 'SureCell',
#         'Smart-seq2;droplet/well-based': 'Smart-seq2',
#         'SCRB-seq;droplet/well-based;inDrop': 'inDrop',
#         'droplet/well-based;inDrop': 'inDrop',
#         'C1/well-based;inDrop': 'C1'
#     }
#     df['platform_new'] = df['platform_new'].replace(replacements)

#     none_list = {
#         'truseq', 'novaseq', 'mrna-seq', 'hiseq2000', 'hiseq2500',
#         'hiseq4000', 'novaseq_sp', 'droplet/well-based', 'geo-seq', 'transplex'
#     }

#     def clean_none(item):
#         parts = [p for p in item.split(';') if p not in none_list]
#         return ';'.join(sorted(set(parts))) if parts else 'none'

#     df['platform_new'] = df['platform_new'].apply(clean_none)

#     smart_combo = {
#         'C1;Smart-seq2', 'Micro-well;Smart-seq2', 'SLAM-Seq;Smart-seq2',
#         'Quartz-seq;Smart-seq2', 'MARS-seq;Smart-seq2', 'SMRT;Smart-seq2',
#         'SCRB-seq;Smart-seq2', 'HyPR-seq;Smart-seq2'
#     }
#     df['platform_new'] = df['platform_new'].replace({k: 'Smart-seq2' for k in smart_combo})

#     # ------------------------------------------------------------------
#     # 7. 整理列并输出
#     # ------------------------------------------------------------------
#     df = df[['gse', 'gsm', 'char_platform', 'platform_key_word', 'platform_new']]
#     df.rename(columns={'platform_new': 'platform_final'}, inplace=True)
#     df = df.drop_duplicates()

#     # 保存 gsm 粒度
#     df.to_csv( platform_gse_gsm_out_csv, sep=',', index=False)

#     # 保存 gse 粒度（仅 gsm 含 GSE 的行）
#     gse_df = df[df['gsm'].str.contains('GSE')]
#     gse_df.to_csv(platform_gse_out, sep=',', index=False)
#     print(df[df['gsm'].str.contains('GSE')].sum())

#     return df, gse_df

